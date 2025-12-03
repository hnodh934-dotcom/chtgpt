#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
إنشاء ملف لجان التظلم والاعتراضات
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# الإعدادات
GREEN = "1B5E4A"
GOLD = "C9A227"
RED = "C0392B"
HEADER_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
GOLD_FILL = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
HEADER_FONT = Font(name='Arial', size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name='Arial', size=16, bold=True, color=GREEN)
SUBTITLE_FONT = Font(name='Arial', size=14, bold=True, color=GOLD)
NORMAL_FONT = Font(name='Arial', size=11)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
RTL_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_data(ws, start_row, data):
    for row_idx, row_data in enumerate(data, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN
            cell.font = NORMAL_FONT

def create_grievance_file():
    wb = Workbook()

    # ================== ورقة 1: نظرة عامة ==================
    ws1 = wb.active
    ws1.title = "نظرة عامة"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:H1')
    ws1['A1'] = "دليل لجان التظلم والاعتراض في المملكة العربية السعودية"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER_ALIGN

    headers = ['م', 'الجهة', 'لجنة التظلم', 'المرجع النظامي', 'مدة التظلم', 'الرسوم', 'المنصة', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=h)
    style_header(ws1, 3, len(headers))

    overview = [
        [1, 'هيئة الغذاء والدواء (SFDA)', 'لجنة النظر في المخالفات', 'المادة 22 - نظام الغذاء', '30 يوم', 'مجاناً', 'sfda.gov.sa', 'درجتان للتظلم'],
        [2, 'وزارة التجارة', 'لجنة الفصل في المخالفات التجارية', 'المادة 23 - نظام مكافحة الغش', '60 يوم', 'مجاناً', 'mc.gov.sa', 'استئناف أمام ديوان المظالم'],
        [3, 'وزارة الموارد البشرية', 'الهيئات العمالية الابتدائية', 'المادة 210 - نظام العمل', '30 يوم', 'مجاناً', 'mol.gov.sa', 'استئناف خلال 30 يوم'],
        [4, 'هيئة الزكاة والضريبة (ZATCA)', 'لجنة الفصل في المخالفات الضريبية', 'المادة 49 - نظام VAT', '30 يوم', 'مجاناً', 'zatca.gov.sa', 'أمانة لجان الفصل'],
        [5, 'الدفاع المدني', 'لجنة التظلمات', 'لائحة السلامة', '15 يوم', 'مجاناً', '998.gov.sa', 'تصعيد للإمارة'],
        [6, 'البلدية/الأمانة', 'لجنة الاستئناف البلدية', 'نظام البلديات', '30 يوم', 'مجاناً', 'balady.gov.sa', 'حسب المنطقة'],
        [7, 'هيئة البيئة', 'لجنة النظر في المخالفات البيئية', 'نظام البيئة', '30 يوم', 'مجاناً', 'mewa.gov.sa', ''],
        [8, 'التأمينات الاجتماعية', 'لجنة تسوية الخلافات', 'نظام التأمينات', '30 يوم', 'مجاناً', 'gosi.gov.sa', 'للمنشآت والأفراد'],
    ]
    add_data(ws1, 4, overview)
    set_widths(ws1, [5, 28, 28, 25, 12, 10, 18, 25])

    # ================== ورقة 2: SFDA ==================
    ws2 = wb.create_sheet("SFDA - الغذاء والدواء")
    ws2.sheet_view.rightToLeft = True

    ws2.merge_cells('A1:G1')
    ws2['A1'] = "إجراءات التظلم أمام هيئة الغذاء والدواء"
    ws2['A1'].font = TITLE_FONT

    ws2['A3'] = "الأساس النظامي:"
    ws2['A3'].font = SUBTITLE_FONT
    ws2['A4'] = "المادة 22 من نظام الغذاء: يجوز لمن صدر بحقه قرار بالعقوبة التظلم أمام اللجنة خلال 30 يوماً"
    ws2.merge_cells('A4:G4')

    headers = ['المرحلة', 'الإجراء', 'المدة', 'المتطلبات', 'النتيجة المتوقعة', 'التصعيد', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws2.cell(row=6, column=col, value=h)
    style_header(ws2, 6, len(headers))

    sfda_steps = [
        ['التظلم الأولي', 'تقديم اعتراض مكتوب للجنة', '30 يوم من القرار', 'صورة القرار + أسباب الاعتراض', 'قبول/رفض/تعديل', 'اللجنة الاستئنافية', ''],
        ['اللجنة الاستئنافية', 'استئناف قرار اللجنة الأولى', '30 يوم من القرار', 'قرار اللجنة الأولى + مبررات', 'قرار نهائي', 'ديوان المظالم', 'قرارها نهائي إدارياً'],
        ['ديوان المظالم', 'الطعن القضائي', '60 يوم', 'جميع المستندات السابقة', 'حكم قضائي', '-', 'آخر درجة'],
    ]
    add_data(ws2, 7, sfda_steps)

    ws2['A11'] = "حالات قبول التظلم:"
    ws2['A11'].font = SUBTITLE_FONT
    cases = [
        "• خطأ في تحديد المخالفة أو تكييفها القانوني",
        "• عدم التناسب بين المخالفة والعقوبة",
        "• وجود ظروف مخففة لم تؤخذ بالاعتبار",
        "• خطأ إجرائي في إصدار القرار",
        "• تقديم أدلة جديدة تؤثر على القرار",
    ]
    for i, case in enumerate(cases, 12):
        ws2[f'A{i}'] = case
        ws2[f'A{i}'].alignment = RTL_ALIGN

    set_widths(ws2, [18, 30, 18, 30, 22, 20, 20])

    # ================== ورقة 3: التجارة ==================
    ws3 = wb.create_sheet("التجارة")
    ws3.sheet_view.rightToLeft = True

    ws3.merge_cells('A1:G1')
    ws3['A1'] = "إجراءات التظلم أمام وزارة التجارة"
    ws3['A1'].font = TITLE_FONT

    ws3['A3'] = "الأساس النظامي:"
    ws3['A3'].font = SUBTITLE_FONT
    ws3['A4'] = "المادة 23 من نظام مكافحة الغش التجاري + المادة 214 من نظام الشركات"
    ws3.merge_cells('A4:G4')

    headers = ['نوع المخالفة', 'اللجنة المختصة', 'مدة التظلم', 'إجراء التظلم', 'جهة الاستئناف', 'المدة النهائية', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws3.cell(row=6, column=col, value=h)
    style_header(ws3, 6, len(headers))

    moci_data = [
        ['غش تجاري', 'لجنة الفصل في مخالفات الغش', '60 يوم', 'طلب عبر المنصة', 'ديوان المظالم', '60 يوم', 'غرامات عالية'],
        ['مخالفات السجل التجاري', 'لجنة السجل التجاري', '30 يوم', 'طلب مكتوب', 'ديوان المظالم', '60 يوم', ''],
        ['مخالفات الشركات', 'لجنة الفصل في منازعات الشركات', '60 يوم', 'طلب إلكتروني', 'ديوان المظالم', '60 يوم', 'نظام الشركات الجديد'],
        ['حماية المستهلك', 'لجنة حماية المستهلك', '30 يوم', 'بلاغ إلكتروني', 'الوزارة', '30 يوم', ''],
        ['التستر التجاري', 'لجنة التستر', '30 يوم', 'طلب رسمي', 'ديوان المظالم', '60 يوم', 'عقوبات مشددة'],
    ]
    add_data(ws3, 7, moci_data)
    set_widths(ws3, [20, 28, 12, 18, 18, 15, 20])

    # ================== ورقة 4: الموارد البشرية ==================
    ws4 = wb.create_sheet("الموارد البشرية")
    ws4.sheet_view.rightToLeft = True

    ws4.merge_cells('A1:G1')
    ws4['A1'] = "إجراءات التظلم العمالي - وزارة الموارد البشرية"
    ws4['A1'].font = TITLE_FONT

    ws4['A3'] = "الأساس النظامي:"
    ws4['A3'].font = SUBTITLE_FONT
    ws4['A4'] = "الباب الرابع عشر من نظام العمل (المواد 210-235) - هيئات تسوية الخلافات العمالية"
    ws4.merge_cells('A4:G4')

    headers = ['المرحلة', 'الجهة', 'الاختصاص', 'المدة', 'الرسوم', 'الاستئناف', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws4.cell(row=6, column=col, value=h)
    style_header(ws4, 6, len(headers))

    hrsd_data = [
        ['التسوية الودية', 'مكتب العمل', 'محاولة الصلح', '21 يوم', 'مجاناً', 'الهيئة الابتدائية', 'إلزامي قبل الدعوى'],
        ['الهيئة الابتدائية', 'هيئة تسوية الخلافات', 'النظر في الدعوى', 'متغير', 'مجاناً', 'الهيئة العليا', 'قرار قابل للاستئناف'],
        ['الهيئة العليا', 'هيئة الاستئناف', 'استئناف الحكم', '30 يوم للاستئناف', 'مجاناً', '-', 'قرار نهائي'],
        ['التنفيذ', 'محكمة التنفيذ', 'تنفيذ الحكم', 'فوري', 'مجاناً', '-', 'بعد اكتساب الصفة النهائية'],
    ]
    add_data(ws4, 7, hrsd_data)

    ws4['A13'] = "مخالفات نظام العمل - التظلم:"
    ws4['A13'].font = SUBTITLE_FONT

    headers2 = ['المخالفة', 'الغرامة', 'مدة التظلم', 'جهة التظلم', 'الإجراء']
    for col, h in enumerate(headers2, 1):
        ws4.cell(row=14, column=col, value=h)
    style_header(ws4, 14, len(headers2))

    violations = [
        ['عدم توثيق العقود', '10,000 ريال/عامل', '30 يوم', 'لجنة مخالفات العمل', 'طلب عبر قوى'],
        ['تأخير الرواتب', '10,000 ريال', '30 يوم', 'لجنة مخالفات العمل', 'طلب عبر قوى'],
        ['مخالفات السعودة', 'إيقاف خدمات', '15 يوم', 'لجنة نطاقات', 'طلب إعادة تقييم'],
        ['مخالفات السلامة', '25,000 ريال', '30 يوم', 'لجنة السلامة المهنية', 'طلب رسمي'],
    ]
    add_data(ws4, 15, violations)
    set_widths(ws4, [22, 25, 18, 22, 22, 18, 20])

    # ================== ورقة 5: ZATCA ==================
    ws5 = wb.create_sheet("ZATCA - الضريبة")
    ws5.sheet_view.rightToLeft = True

    ws5.merge_cells('A1:G1')
    ws5['A1'] = "إجراءات التظلم الضريبي - هيئة الزكاة والضريبة والجمارك"
    ws5['A1'].font = TITLE_FONT

    ws5['A3'] = "الأساس النظامي:"
    ws5['A3'].font = SUBTITLE_FONT
    ws5['A4'] = "المادة 49 من نظام ضريبة القيمة المضافة + نظام الزكاة + نظام الجمارك الموحد"
    ws5.merge_cells('A4:G4')

    headers = ['نوع التظلم', 'الجهة', 'المدة', 'المتطلبات', 'الرسوم', 'النتيجة', 'التصعيد']
    for col, h in enumerate(headers, 1):
        ws5.cell(row=6, column=col, value=h)
    style_header(ws5, 6, len(headers))

    zatca_data = [
        ['اعتراض على الربط الضريبي', 'الهيئة - إدارة الاعتراضات', '60 يوم من الإشعار', 'خطاب اعتراض + مستندات', 'مجاناً', 'قبول/رفض/تعديل', 'لجنة الفصل'],
        ['لجنة الفصل الابتدائية', 'أمانة لجان الفصل', '30 يوم من رد الهيئة', 'قرار الهيئة + اعتراض', 'مجاناً', 'حكم ابتدائي', 'اللجنة الاستئنافية'],
        ['اللجنة الاستئنافية', 'لجنة الاستئناف', '30 يوم من الحكم', 'الحكم الابتدائي + استئناف', 'مجاناً', 'حكم نهائي', 'ديوان المظالم'],
        ['الطعن القضائي', 'ديوان المظالم', '60 يوم', 'جميع المستندات', 'مجاناً', 'حكم قضائي', '-'],
        ['تقسيط الغرامات', 'الهيئة', 'قبل التنفيذ', 'طلب + ضمان', 'مجاناً', 'جدول سداد', '-'],
        ['الإعفاء من الغرامات', 'الهيئة', 'حسب المبادرة', 'طلب + مبررات', 'مجاناً', 'إعفاء جزئي/كلي', '-'],
    ]
    add_data(ws5, 7, zatca_data)

    ws5['A15'] = "مبادرات الإعفاء من الغرامات:"
    ws5['A15'].font = SUBTITLE_FONT
    exemptions = [
        "• مبادرة إلغاء الغرامات: للمنشآت المتأخرة عن التسجيل والإقرار (تفعل دورياً)",
        "• تقسيط المستحقات: حتى 24 شهر بدون فوائد",
        "• الإعفاء للظروف القاهرة: كوارث، إفلاس، ظروف استثنائية",
    ]
    for i, ex in enumerate(exemptions, 16):
        ws5[f'A{i}'] = ex
        ws5[f'A{i}'].alignment = RTL_ALIGN

    set_widths(ws5, [25, 25, 20, 28, 12, 20, 18])

    # ================== ورقة 6: الدفاع المدني ==================
    ws6 = wb.create_sheet("الدفاع المدني")
    ws6.sheet_view.rightToLeft = True

    ws6.merge_cells('A1:F1')
    ws6['A1'] = "إجراءات التظلم - الدفاع المدني"
    ws6['A1'].font = TITLE_FONT

    headers = ['نوع المخالفة', 'الغرامة', 'جهة التظلم', 'المدة', 'الإجراء', 'التصعيد']
    for col, h in enumerate(headers, 1):
        ws6.cell(row=3, column=col, value=h)
    style_header(ws6, 3, len(headers))

    cd_data = [
        ['عدم توفر وسائل السلامة', '20,000 - 50,000', 'مدير الإدارة', '15 يوم', 'خطاب رسمي', 'مدير المنطقة'],
        ['انتهاء صلاحية الطفايات', '10,000 - 20,000', 'مدير الإدارة', '15 يوم', 'خطاب + إثبات الإصلاح', 'مدير المنطقة'],
        ['مخالفات الإخلاء', '30,000 - 50,000', 'مدير الإدارة', '15 يوم', 'خطاب رسمي', 'الإمارة'],
        ['قرار الإغلاق', 'إغلاق المنشأة', 'مدير المنطقة', '15 يوم', 'تظلم + خطة تصحيح', 'الإمارة'],
        ['الإغلاق النهائي', 'إغلاق دائم', 'الإمارة', '30 يوم', 'طلب إعادة نظر', 'ديوان المظالم'],
    ]
    add_data(ws6, 4, cd_data)
    set_widths(ws6, [25, 18, 18, 12, 25, 18])

    # ================== ورقة 7: البلدية ==================
    ws7 = wb.create_sheet("البلدية")
    ws7.sheet_view.rightToLeft = True

    ws7.merge_cells('A1:F1')
    ws7['A1'] = "إجراءات التظلم - الأمانات والبلديات"
    ws7['A1'].font = TITLE_FONT

    ws7['A3'] = "الأساس النظامي: نظام البلديات والقرى + لائحة الجزاءات البلدية"
    ws7['A3'].font = SUBTITLE_FONT
    ws7.merge_cells('A3:F3')

    headers = ['المخالفة', 'الغرامة', 'جهة التظلم الأولى', 'جهة الاستئناف', 'المدة', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws7.cell(row=5, column=col, value=h)
    style_header(ws7, 5, len(headers))

    mun_data = [
        ['مخالفات الرخصة', '1,000 - 10,000', 'رئيس البلدية', 'لجنة الاستئناف', '30 يوم', 'بلدي'],
        ['مخالفات النظافة', '500 - 5,000', 'رئيس البلدية', 'أمين المنطقة', '30 يوم', ''],
        ['مخالفات البناء', '10,000 - 100,000', 'لجنة المخالفات', 'لجنة الاستئناف', '30 يوم', 'قد تشمل الإزالة'],
        ['مخالفات اللوحات', '1,000 - 5,000', 'رئيس البلدية', 'أمين المنطقة', '15 يوم', ''],
        ['قرار الإغلاق', 'إغلاق', 'أمين المنطقة', 'الوزير', '30 يوم', 'طلب رفع الإغلاق'],
    ]
    add_data(ws7, 6, mun_data)
    set_widths(ws7, [22, 18, 22, 20, 12, 22])

    # ================== ورقة 8: نموذج تظلم ==================
    ws8 = wb.create_sheet("نموذج التظلم")
    ws8.sheet_view.rightToLeft = True

    ws8.merge_cells('A1:D1')
    ws8['A1'] = "سجل متابعة التظلمات والاعتراضات"
    ws8['A1'].font = TITLE_FONT

    headers = ['م', 'رقم المخالفة', 'الجهة', 'تاريخ المخالفة', 'الغرامة', 'تاريخ التظلم',
               'جهة التظلم', 'حالة التظلم', 'نتيجة التظلم', 'الغرامة بعد التظلم', 'الوفر', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws8.cell(row=3, column=col, value=h)
    style_header(ws8, 3, len(headers))

    # صفوف فارغة للتعبئة مع معادلات
    for row in range(4, 20):
        for col in range(1, 13):
            cell = ws8.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
        # معادلة الوفر
        ws8.cell(row=row, column=11, value=f'=IF(AND(E{row}<>"",J{row}<>""),E{row}-J{row},"")')

    # Data Validation للحالة
    status_dv = DataValidation(type="list",
        formula1='"قيد المراجعة,مقبول,مرفوض,معدل,مستأنف,منتهي"')
    ws8.add_data_validation(status_dv)
    status_dv.add('H4:H100')

    # Data Validation للجهة
    agency_dv = DataValidation(type="list",
        formula1='"SFDA,التجارة,الموارد البشرية,ZATCA,الدفاع المدني,البلدية,البيئة,التأمينات"')
    ws8.add_data_validation(agency_dv)
    agency_dv.add('C4:C100')

    # صف الإجمالي
    ws8['A21'] = 'الإجمالي'
    ws8['A21'].font = Font(bold=True)
    ws8['E21'] = '=SUM(E4:E20)'
    ws8['J21'] = '=SUM(J4:J20)'
    ws8['K21'] = '=SUM(K4:K20)'

    set_widths(ws8, [5, 15, 15, 12, 12, 12, 15, 12, 15, 15, 12, 20])

    # ================== ورقة 9: الإحصائيات ==================
    ws9 = wb.create_sheet("إحصائيات التظلم")
    ws9.sheet_view.rightToLeft = True

    ws9.merge_cells('A1:E1')
    ws9['A1'] = "إحصائيات نجاح التظلمات (بيانات مرجعية)"
    ws9['A1'].font = TITLE_FONT

    headers = ['الجهة', 'نسبة القبول', 'متوسط التخفيض', 'متوسط المدة', 'توصية']
    for col, h in enumerate(headers, 1):
        ws9.cell(row=3, column=col, value=h)
    style_header(ws9, 3, len(headers))

    stats = [
        ['SFDA', '35%', '40%', '45 يوم', 'التظلم مجدي للمخالفات الكبيرة'],
        ['التجارة', '40%', '30%', '60 يوم', 'التظلم مجدي'],
        ['الموارد البشرية', '50%', '50%', '90 يوم', 'التسوية الودية أفضل'],
        ['ZATCA', '45%', '25%', '120 يوم', 'المبادرات أفضل من التظلم'],
        ['الدفاع المدني', '30%', '20%', '30 يوم', 'الإصلاح أسرع من التظلم'],
        ['البلدية', '35%', '30%', '45 يوم', 'حسب نوع المخالفة'],
    ]
    add_data(ws9, 4, stats)

    ws9['A12'] = "ملاحظة: هذه نسب تقريبية بناءً على الخبرة العملية وليست رسمية"
    ws9['A12'].font = Font(italic=True, color="888888")

    set_widths(ws9, [20, 15, 18, 15, 35])

    # حفظ الملف
    wb.save('grievance_committees.xlsx')
    print("✓ تم إنشاء: grievance_committees.xlsx")

if __name__ == "__main__":
    os.chdir('/home/user/chtgpt/food_factory_compliance/02_regulatory_framework')
    create_grievance_file()
