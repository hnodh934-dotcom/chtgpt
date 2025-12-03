#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
إنشاء ملفات الإطار التنظيمي
02_regulatory_framework
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
import os

# الألوان والتنسيقات
GREEN = "1B5E4A"
GOLD = "C9A227"
HEADER_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
HEADER_FONT = Font(name='Arial', size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name='Arial', size=16, bold=True, color=GREEN)
NORMAL_FONT = Font(name='Arial', size=11)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
RTL_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

def set_column_widths(ws, widths):
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_sfda_requirements():
    """متطلبات هيئة الغذاء والدواء"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "متطلبات SFDA"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:H1')
    ws1['A1'] = "متطلبات هيئة الغذاء والدواء السعودية (SFDA)"
    ws1['A1'].font = TITLE_FONT

    headers = ['م', 'المتطلب', 'المرجع النظامي', 'الوصف التفصيلي', 'الحالة', 'الأولوية', 'المدة', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    requirements = [
        [1, 'ترخيص منشأة غذائية', 'المادة 8 - نظام الغذاء', 'الحصول على ترخيص ساري من SFDA لممارسة النشاط', '', 'حرجة', '30 يوم', ''],
        [2, 'نظام HACCP', 'المادة 15 - نظام الغذاء', 'تطبيق نظام تحليل المخاطر ونقاط التحكم الحرجة', '', 'حرجة', '90 يوم', ''],
        [3, 'شهادة صحية للعاملين', 'المادة 10 - نظام الغذاء', 'فحوصات طبية دورية لجميع العاملين في إنتاج الغذاء', '', 'عالية', '14 يوم', 'كل 6 أشهر'],
        [4, 'سجل التتبع', 'المادة 18 - نظام الغذاء', 'نظام تتبع المنتجات من المورد إلى المستهلك', '', 'عالية', '30 يوم', ''],
        [5, 'شروط التخزين', 'GSO 839', 'درجات الحرارة والرطوبة المناسبة للمنتجات', '', 'عالية', '14 يوم', ''],
        [6, 'البطاقة الغذائية', 'GSO 9', 'بطاقة غذائية مطابقة للمواصفات الخليجية', '', 'عالية', '30 يوم', ''],
        [7, 'إدارة الاستدعاء', 'المادة 20 - نظام الغذاء', 'إجراء موثق لاستدعاء المنتجات عند الحاجة', '', 'متوسطة', '14 يوم', ''],
        [8, 'مراقبة الآفات', 'GSO 21', 'برنامج متكامل لمكافحة الآفات مع شركة مرخصة', '', 'عالية', '7 أيام', 'عقد سنوي'],
        [9, 'جودة المياه', 'SASO 701', 'تحليل دوري لمياه الإنتاج والتنظيف', '', 'عالية', '7 أيام', 'شهري'],
        [10, 'التدريب', 'المادة 11 - نظام الغذاء', 'برنامج تدريب على سلامة الغذاء لجميع العاملين', '', 'عالية', '30 يوم', ''],
    ]

    for row_idx, req in enumerate(requirements, 4):
        for col_idx, value in enumerate(req, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN
            cell.font = NORMAL_FONT

    # Data Validation
    status_dv = DataValidation(type="list", formula1='"مطبق,جزئي,غير مطبق,قيد التنفيذ"')
    ws1.add_data_validation(status_dv)
    status_dv.add('E4:E100')

    priority_dv = DataValidation(type="list", formula1='"حرجة,عالية,متوسطة,منخفضة"')
    ws1.add_data_validation(priority_dv)
    priority_dv.add('F4:F100')

    set_column_widths(ws1, [5, 25, 22, 40, 12, 12, 12, 25])

    # ورقة العقوبات
    ws2 = wb.create_sheet("العقوبات")
    ws2.sheet_view.rightToLeft = True

    ws2['A1'] = "عقوبات مخالفات نظام الغذاء"
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells('A1:E1')

    penalty_headers = ['المخالفة', 'العقوبة الأولى', 'التكرار', 'الإغلاق', 'المرجع']
    for col, header in enumerate(penalty_headers, 1):
        ws2.cell(row=3, column=col, value=header)
    style_header_row(ws2, 3, len(penalty_headers))

    penalties = [
        ['عدم الترخيص', '50,000 - 100,000 ريال', 'الإغلاق النهائي', 'نعم', 'المادة 23'],
        ['منتج منتهي الصلاحية', '100,000 - 500,000 ريال', 'الإغلاق + السجن', 'نعم', 'المادة 24'],
        ['غش تجاري', '500,000 - 1,000,000 ريال', 'السجن 3 سنوات', 'نعم', 'المادة 25'],
        ['عدم تطبيق HACCP', '50,000 - 150,000 ريال', 'الإيقاف', 'نعم', 'المادة 23'],
        ['عدم توفر الشهادات الصحية', '25,000 - 50,000 ريال', '100,000 ريال', 'لا', 'المادة 23'],
    ]

    for row_idx, pen in enumerate(penalties, 4):
        for col_idx, value in enumerate(pen, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN

    set_column_widths(ws2, [25, 25, 25, 10, 12])

    wb.save('02_regulatory_framework/sfda_requirements.xlsx')
    print("✓ تم إنشاء: sfda_requirements.xlsx")

def create_saso_standards():
    """مواصفات هيئة المواصفات السعودية"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "مواصفات SASO"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:G1')
    ws1['A1'] = "المواصفات القياسية السعودية والخليجية للأغذية"
    ws1['A1'].font = TITLE_FONT

    headers = ['م', 'رقم المواصفة', 'اسم المواصفة', 'النطاق', 'إلزامية', 'الحالة', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    standards = [
        [1, 'GSO 9', 'بطاقات المنتجات الغذائية المعبأة', 'جميع المنتجات المعبأة', 'نعم', '', 'أساسية للتصدير'],
        [2, 'GSO 21', 'الشروط الصحية في مصانع الأغذية', 'جميع المصانع', 'نعم', '', ''],
        [3, 'GSO 839', 'تخزين المواد الغذائية المبردة', 'منتجات مبردة', 'نعم', '', ''],
        [4, 'GSO 1016', 'شروط النظافة العامة للأغذية', 'جميع المنشآت', 'نعم', '', 'GMP'],
        [5, 'GSO 150', 'المواد الملامسة للأغذية', 'مواد التعبئة', 'نعم', '', ''],
        [6, 'GSO 2233', 'إدارة سلامة الغذاء - HACCP', 'مطلوب', 'نعم', '', ''],
        [7, 'SASO 701', 'مواصفات مياه الشرب', 'مياه الإنتاج', 'نعم', '', ''],
        [8, 'GSO CAC/RCP 1', 'الممارسات الصحية الجيدة', 'جميع المنشآت', 'نعم', '', 'GHP'],
        [9, 'GSO 654', 'المضافات الغذائية المسموحة', 'منتجات بمضافات', 'نعم', '', ''],
        [10, 'GSO 2500', 'الأغذية الحلال', 'جميع المنتجات', 'نعم', '', 'للتصدير'],
    ]

    for row_idx, std in enumerate(standards, 4):
        for col_idx, value in enumerate(std, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN

    status_dv = DataValidation(type="list", formula1='"مطبق,جزئي,غير مطبق"')
    ws1.add_data_validation(status_dv)
    status_dv.add('F4:F100')

    set_column_widths(ws1, [5, 15, 35, 25, 10, 12, 25])
    wb.save('02_regulatory_framework/saso_standards.xlsx')
    print("✓ تم إنشاء: saso_standards.xlsx")

def create_moci_requirements():
    """متطلبات وزارة التجارة"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "متطلبات التجارة"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:G1')
    ws1['A1'] = "متطلبات وزارة التجارة - نظام الشركات والسجل التجاري"
    ws1['A1'].font = TITLE_FONT

    headers = ['م', 'المتطلب', 'المرجع النظامي', 'المدة', 'الرسوم', 'الحالة', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    requirements = [
        [1, 'السجل التجاري', 'نظام السجل التجاري', 'سنوي', '1,200 ريال', '', 'أساسي لجميع التراخيص'],
        [2, 'عضوية الغرفة التجارية', 'نظام الغرف التجارية', 'سنوي', '500-2000 ريال', '', 'حسب حجم المنشأة'],
        [3, 'رخصة الاستيراد', 'نظام الجمارك', 'سنوي', '2,000 ريال', '', 'للمواد الخام'],
        [4, 'رخصة التصدير', 'نظام الصادرات', 'سنوي', '2,000 ريال', '', 'للتصدير'],
        [5, 'شهادة المنشأ', 'نظام التصدير', 'لكل شحنة', '100 ريال', '', 'للتصدير'],
        [6, 'العلامة التجارية', 'نظام العلامات التجارية', '10 سنوات', '3,000 ريال', '', 'اختياري'],
        [7, 'حماية المستهلك', 'نظام حماية المستهلك', 'مستمر', '-', '', 'الامتثال مطلوب'],
        [8, 'مكافحة الغش', 'نظام مكافحة الغش', 'مستمر', '-', '', 'عقوبات مشددة'],
    ]

    for row_idx, req in enumerate(requirements, 4):
        for col_idx, value in enumerate(req, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN

    status_dv = DataValidation(type="list", formula1='"ساري,منتهي,قيد التجديد,غير مطبق"')
    ws1.add_data_validation(status_dv)
    status_dv.add('F4:F100')

    set_column_widths(ws1, [5, 25, 25, 12, 15, 12, 25])
    wb.save('02_regulatory_framework/moci_requirements.xlsx')
    print("✓ تم إنشاء: moci_requirements.xlsx")

def create_hrsd_requirements():
    """متطلبات الموارد البشرية"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "نظام العمل"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:G1')
    ws1['A1'] = "متطلبات وزارة الموارد البشرية والتنمية الاجتماعية"
    ws1['A1'].font = TITLE_FONT

    headers = ['م', 'المتطلب', 'المادة النظامية', 'الوصف', 'العقوبة', 'الحالة', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    requirements = [
        [1, 'عقود العمل الموثقة', 'المادة 51', 'توثيق جميع عقود العمل إلكترونياً', '10,000 ريال/عامل', '', ''],
        [2, 'نطاقات (السعودة)', 'المادة 26', 'تحقيق نسبة التوطين المطلوبة', 'إيقاف الخدمات', '', 'حسب النشاط'],
        [3, 'حماية الأجور', 'المادة 90', 'صرف الرواتب عبر النظام البنكي', '10,000 ريال', '', 'شهري'],
        [4, 'التأمينات الاجتماعية', 'نظام التأمينات', 'تسجيل جميع العاملين', 'غرامات تراكمية', '', ''],
        [5, 'السلامة المهنية', 'المادة 121', 'توفير بيئة عمل آمنة', '25,000 ريال', '', ''],
        [6, 'ساعات العمل', 'المادة 98', 'الالتزام بساعات العمل القانونية', '10,000 ريال', '', '8 ساعات/يوم'],
        [7, 'الإجازات', 'المادة 109', 'منح الإجازات السنوية', '5,000 ريال', '', '21-30 يوم'],
        [8, 'نهاية الخدمة', 'المادة 84', 'احتساب مكافأة نهاية الخدمة', 'التعويض + غرامة', '', ''],
        [9, 'التدريب', 'المادة 42', 'برامج تدريب للسعوديين', '-', '', 'اختياري'],
        [10, 'الفحص الطبي', 'المادة 10', 'فحص طبي قبل التوظيف', '-', '', 'للأغذية: إلزامي'],
    ]

    for row_idx, req in enumerate(requirements, 4):
        for col_idx, value in enumerate(req, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN

    status_dv = DataValidation(type="list", formula1='"ممتثل,جزئي,غير ممتثل"')
    ws1.add_data_validation(status_dv)
    status_dv.add('F4:F100')

    set_column_widths(ws1, [5, 22, 15, 35, 18, 12, 20])
    wb.save('02_regulatory_framework/hrsd_requirements.xlsx')
    print("✓ تم إنشاء: hrsd_requirements.xlsx")

def create_zatca_requirements():
    """متطلبات هيئة الزكاة والضريبة"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "متطلبات ZATCA"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:G1')
    ws1['A1'] = "متطلبات هيئة الزكاة والضريبة والجمارك (ZATCA)"
    ws1['A1'].font = TITLE_FONT

    headers = ['م', 'المتطلب', 'النظام', 'الموعد', 'العقوبة', 'الحالة', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    requirements = [
        [1, 'التسجيل في ضريبة القيمة المضافة', 'نظام VAT', 'إلزامي > 375,000', '50% من الضريبة', '', ''],
        [2, 'الإقرار الضريبي', 'نظام VAT', 'ربع سنوي', '5-25% غرامة', '', ''],
        [3, 'الفاتورة الإلكترونية', 'لائحة الفوترة', 'إلزامي', '50,000 ريال', '', 'FATOORAH'],
        [4, 'إقرار الزكاة', 'نظام الزكاة', 'سنوي', '1% شهرياً', '', ''],
        [5, 'شهادة الزكاة', 'نظام الزكاة', 'للتجديدات', 'إيقاف الخدمات', '', ''],
        [6, 'الاستقطاع الضريبي', 'نظام الاستقطاع', 'شهري', '1% + 5,000', '', 'للمدفوعات الخارجية'],
        [7, 'الجمارك', 'نظام الجمارك', 'لكل شحنة', '200% + مصادرة', '', 'استيراد/تصدير'],
        [8, 'السجلات المحاسبية', 'نظام VAT', 'مستمر', '50,000 ريال', '', '5 سنوات حفظ'],
    ]

    for row_idx, req in enumerate(requirements, 4):
        for col_idx, value in enumerate(req, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN

    status_dv = DataValidation(type="list", formula1='"ممتثل,متأخر,غير مسجل"')
    ws1.add_data_validation(status_dv)
    status_dv.add('F4:F100')

    set_column_widths(ws1, [5, 30, 15, 18, 18, 12, 25])
    wb.save('02_regulatory_framework/zatca_requirements.xlsx')
    print("✓ تم إنشاء: zatca_requirements.xlsx")

def create_civil_defense():
    """اشتراطات الدفاع المدني"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "السلامة"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:G1')
    ws1['A1'] = "اشتراطات الدفاع المدني والسلامة"
    ws1['A1'].font = TITLE_FONT

    headers = ['م', 'المتطلب', 'المواصفة', 'التكرار', 'العقوبة', 'الحالة', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    requirements = [
        [1, 'رخصة الدفاع المدني', 'لائحة السلامة', 'سنوي', 'إغلاق المنشأة', '', 'أساسي'],
        [2, 'طفايات الحريق', 'SASO 1927', 'فحص سنوي', '20,000 ريال', '', 'حسب المساحة'],
        [3, 'كاشفات الدخان', 'NFPA 72', 'فحص شهري', '15,000 ريال', '', ''],
        [4, 'مخارج الطوارئ', 'كود البناء', 'معاينة', '50,000 + إغلاق', '', 'لافتات مضيئة'],
        [5, 'شبكة الرش الآلي', 'NFPA 13', 'إن وجد', '30,000 ريال', '', 'حسب الحجم'],
        [6, 'خزان مياه الحريق', 'لائحة السلامة', 'سنوي', '25,000 ريال', '', ''],
        [7, 'خطة الإخلاء', 'لائحة السلامة', 'تدريب سنوي', '10,000 ريال', '', 'توثيق'],
        [8, 'تصريح المواد الخطرة', 'لائحة المواد', 'سنوي', '50,000 ريال', '', 'إن وجد'],
        [9, 'سجل السلامة', 'لائحة السلامة', 'مستمر', '5,000 ريال', '', ''],
        [10, 'مسؤول السلامة', 'لائحة السلامة', 'معين', '10,000 ريال', '', 'تدريب معتمد'],
    ]

    for row_idx, req in enumerate(requirements, 4):
        for col_idx, value in enumerate(req, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN

    status_dv = DataValidation(type="list", formula1='"متوفر,غير متوفر,منتهي"')
    ws1.add_data_validation(status_dv)
    status_dv.add('F4:F100')

    set_column_widths(ws1, [5, 25, 15, 12, 18, 12, 25])
    wb.save('02_regulatory_framework/civil_defense.xlsx')
    print("✓ تم إنشاء: civil_defense.xlsx")

def create_environmental_permits():
    """التصاريح البيئية"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "البيئة"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:G1')
    ws1['A1'] = "متطلبات المركز الوطني للرقابة البيئية"
    ws1['A1'].font = TITLE_FONT

    headers = ['م', 'المتطلب', 'المرجع', 'التصنيف', 'الموعد', 'الحالة', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    requirements = [
        [1, 'الرخصة البيئية', 'نظام البيئة', 'إلزامي', 'سنوي', '', 'قبل التشغيل'],
        [2, 'تقييم الأثر البيئي', 'نظام البيئة', 'الفئة 2', 'مرة واحدة', '', 'للمصانع'],
        [3, 'إدارة النفايات', 'لائحة النفايات', 'إلزامي', 'عقد سنوي', '', 'شركة مرخصة'],
        [4, 'معالجة مياه الصرف', 'معايير التصريف', 'إلزامي', 'فحص ربعي', '', ''],
        [5, 'الانبعاثات الهوائية', 'معايير الهواء', 'إن وجد', 'فحص سنوي', '', 'المداخن'],
        [6, 'الضوضاء', 'معايير الضوضاء', 'إلزامي', 'قياس سنوي', '', '<70 ديسيبل'],
        [7, 'تخزين المواد الكيميائية', 'لائحة المواد', 'إلزامي', 'معاينة', '', 'MSDS'],
        [8, 'خطة الطوارئ البيئية', 'نظام البيئة', 'إلزامي', 'تحديث سنوي', '', ''],
    ]

    for row_idx, req in enumerate(requirements, 4):
        for col_idx, value in enumerate(req, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN

    status_dv = DataValidation(type="list", formula1='"ساري,منتهي,غير موجود"')
    ws1.add_data_validation(status_dv)
    status_dv.add('F4:F100')

    set_column_widths(ws1, [5, 25, 18, 12, 15, 12, 25])
    wb.save('02_regulatory_framework/environmental_permits.xlsx')
    print("✓ تم إنشاء: environmental_permits.xlsx")

def create_municipality_requirements():
    """اشتراطات البلدية"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "البلدية"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:G1')
    ws1['A1'] = "اشتراطات الأمانة والبلدية"
    ws1['A1'].font = TITLE_FONT

    headers = ['م', 'المتطلب', 'المرجع', 'المدة', 'الرسوم', 'الحالة', 'ملاحظات']
    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    requirements = [
        [1, 'رخصة البلدية', 'لائحة الرخص', 'سنوي', '3,000-10,000', '', 'حسب النشاط'],
        [2, 'شهادة مطابقة المبنى', 'كود البناء', 'مرة واحدة', '2,000 ريال', '', ''],
        [3, 'لوحة المنشأة', 'لائحة اللوحات', 'سنوي', '500 ريال', '', 'مواصفات محددة'],
        [4, 'عقد نظافة', 'لائحة النظافة', 'سنوي', 'حسب المساحة', '', 'شركة مرخصة'],
        [5, 'صحة البيئة', 'لائحة الصحة', 'معاينة', '-', '', 'فجائية'],
        [6, 'موقف سيارات', 'لائحة المواقف', 'ثابت', '-', '', 'حسب المساحة'],
        [7, 'واجهة المبنى', 'لائحة الواجهات', 'صيانة', '-', '', ''],
        [8, 'تصريح الإعلانات', 'لائحة الإعلانات', 'سنوي', '1,000 ريال', '', 'لكل إعلان'],
    ]

    for row_idx, req in enumerate(requirements, 4):
        for col_idx, value in enumerate(req, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN

    status_dv = DataValidation(type="list", formula1='"ساري,منتهي,غير موجود"')
    ws1.add_data_validation(status_dv)
    status_dv.add('F4:F100')

    set_column_widths(ws1, [5, 22, 18, 12, 18, 12, 25])
    wb.save('02_regulatory_framework/municipality_requirements.xlsx')
    print("✓ تم إنشاء: municipality_requirements.xlsx")

if __name__ == "__main__":
    os.chdir('/home/user/chtgpt/food_factory_compliance')
    print("جاري إنشاء ملفات الإطار التنظيمي...")
    create_sfda_requirements()
    create_saso_standards()
    create_moci_requirements()
    create_hrsd_requirements()
    create_zatca_requirements()
    create_civil_defense()
    create_environmental_permits()
    create_municipality_requirements()
    print("\n✓ تم إنشاء جميع ملفات 02_regulatory_framework بنجاح!")
