#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
دمج جميع ملفات Excel في كل مجلد إلى ملف واحد
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
import os
import shutil

# الألوان
GREEN = "1B5E4A"
GOLD = "C9A227"
TITLE_FONT = Font(name='Arial', size=18, bold=True, color=GREEN)

def copy_sheet(source_ws, target_ws):
    """نسخ محتوى ورقة إلى ورقة أخرى"""
    # نسخ أبعاد الأعمدة
    for col_letter, col_dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = col_dim.width

    # نسخ أبعاد الصفوف
    for row_num, row_dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_num].height = row_dim.height

    # نسخ الخلايا
    for row in source_ws.rows:
        for cell in row:
            new_cell = target_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # نسخ الخلايا المدمجة
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    # نسخ Data Validation
    for dv in source_ws.data_validations.dataValidation:
        target_ws.add_data_validation(dv)

    # نسخ Conditional Formatting
    for cf_range, rules in source_ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            target_ws.conditional_formatting.add(str(cf_range), rule)

    # نسخ إعداد RTL
    target_ws.sheet_view.rightToLeft = source_ws.sheet_view.rightToLeft

def merge_folder(folder_path, output_name, folder_title):
    """دمج جميع ملفات Excel في مجلد واحد"""
    print(f"\n📁 {folder_title}")

    # جمع جميع ملفات Excel
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]

    if not excel_files:
        print("   لا توجد ملفات Excel")
        return

    # إنشاء ملف جديد
    merged_wb = Workbook()
    merged_wb.remove(merged_wb.active)  # حذف الورقة الافتراضية

    sheet_count = 0

    for excel_file in sorted(excel_files):
        file_path = os.path.join(folder_path, excel_file)
        try:
            source_wb = load_workbook(file_path)

            for sheet_name in source_wb.sheetnames:
                source_ws = source_wb[sheet_name]

                # اسم الورقة الجديد (مختصر إذا كان طويلاً)
                # Excel يحد الاسم بـ 31 حرف
                base_name = excel_file.replace('.xlsx', '').replace('_', ' ')
                if len(source_wb.sheetnames) > 1:
                    new_name = f"{base_name[:20]}-{sheet_name[:10]}"
                else:
                    new_name = base_name[:31]

                # التأكد من عدم تكرار الاسم
                counter = 1
                original_name = new_name
                while new_name in merged_wb.sheetnames:
                    new_name = f"{original_name[:28]}{counter}"
                    counter += 1

                # إنشاء ورقة جديدة ونسخ المحتوى
                target_ws = merged_wb.create_sheet(title=new_name[:31])
                copy_sheet(source_ws, target_ws)
                sheet_count += 1
                print(f"   ✓ {new_name}")

            source_wb.close()

        except Exception as e:
            print(f"   ✗ خطأ في {excel_file}: {e}")

    # حفظ الملف المدمج
    output_path = os.path.join(folder_path, output_name)
    merged_wb.save(output_path)
    print(f"   📊 تم إنشاء: {output_name} ({sheet_count} ورقة)")

    return sheet_count

def cleanup_old_files(folder_path, keep_file):
    """حذف الملفات القديمة وإبقاء الملف المدمج فقط"""
    for f in os.listdir(folder_path):
        if f.endswith('.xlsx') and f != keep_file:
            os.remove(os.path.join(folder_path, f))
        elif f.endswith('.py'):
            os.remove(os.path.join(folder_path, f))

# التكوين - اسم كل ملف مدمج
folders_config = [
    ('01_current_status', '01_الحالة_الراهنة.xlsx', 'الحالة الراهنة'),
    ('02_regulatory_framework', '02_الإطار_التنظيمي.xlsx', 'الإطار التنظيمي'),
    ('03_compliance_roadmap', '03_خارطة_الامتثال.xlsx', 'خارطة الامتثال'),
    ('04_certifications', '04_الشهادات_الدولية.xlsx', 'الشهادات الدولية'),
    ('05_export_readiness', '05_جاهزية_التصدير.xlsx', 'جاهزية التصدير'),
    ('06_risk_management', '06_إدارة_المخاطر.xlsx', 'إدارة المخاطر'),
    ('07_kpis_dashboard', '07_لوحة_المؤشرات.xlsx', 'لوحة المؤشرات'),
    ('08_documentation', '08_التوثيق.xlsx', 'التوثيق'),
    ('09_financial_planning', '09_التخطيط_المالي.xlsx', 'التخطيط المالي'),
    ('10_checklists', '10_قوائم_الفحص.xlsx', 'قوائم الفحص'),
    ('11_master_files', '11_الملفات_الرئيسية.xlsx', 'الملفات الرئيسية'),
]

if __name__ == "__main__":
    os.chdir('/home/user/chtgpt/food_factory_compliance')

    print("="*60)
    print("دمج ملفات Excel في ملفات موحدة")
    print("="*60)

    total_sheets = 0

    for folder, output, title in folders_config:
        if os.path.exists(folder):
            sheets = merge_folder(folder, output, title)
            if sheets:
                total_sheets += sheets

    print("\n" + "="*60)
    print(f"✓ تم الدمج بنجاح!")
    print(f"  إجمالي الأوراق: {total_sheets}")
    print(f"  إجمالي الملفات: {len(folders_config)}")
    print("="*60)

    # تنظيف الملفات القديمة
    print("\n🧹 تنظيف الملفات القديمة...")
    for folder, output, title in folders_config:
        if os.path.exists(folder):
            cleanup_old_files(folder, output)
            print(f"   ✓ {folder}")

    print("\n✓ تم التنظيف!")
