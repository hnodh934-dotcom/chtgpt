#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
إنشاء ملفات الحالة الراهنة للمصنع
01_current_status
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
import os

# الألوان
GREEN = "1B5E4A"
GOLD = "C9A227"
RED = "C0392B"
ORANGE = "E67E22"
LIGHT_GREEN = "27AE60"
HEADER_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
GOLD_FILL = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
RED_FILL = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
LIGHT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# الخطوط
HEADER_FONT = Font(name='Arial', size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name='Arial', size=16, bold=True, color=GREEN)
NORMAL_FONT = Font(name='Arial', size=11)

# الحدود
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# المحاذاة RTL
RTL_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header_row(ws, row, cols):
    """تنسيق صف الرأس"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

def set_column_widths(ws, widths):
    """تعيين عرض الأعمدة"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

def create_violation_register():
    """إنشاء سجل المخالفات الحالية"""
    wb = Workbook()

    # ورقة 1: سجل المخالفات
    ws1 = wb.active
    ws1.title = "سجل المخالفات"
    ws1.sheet_view.rightToLeft = True

    # العنوان
    ws1.merge_cells('A1:J1')
    ws1['A1'] = "سجل المخالفات والملاحظات الرقابية"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER_ALIGN

    # رؤوس الأعمدة
    headers = ['م', 'رقم المخالفة', 'تاريخ الرصد', 'الجهة الرقابية', 'نوع المخالفة',
               'وصف المخالفة', 'المادة النظامية', 'الغرامة (ريال)', 'الحالة', 'تاريخ التصحيح']

    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    # بيانات نموذجية للمخالفات
    violations = [
        [1, 'VIO-2024-001', '2024-01-15', 'هيئة الغذاء والدواء', 'سلامة غذائية',
         'عدم وجود نظام HACCP موثق', 'المادة 15 - نظام الغذاء', 50000, 'قائمة', ''],
        [2, 'VIO-2024-002', '2024-01-15', 'هيئة الغذاء والدواء', 'نظافة',
         'قصور في معايير النظافة العامة', 'المادة 12 - نظام الغذاء', 25000, 'قائمة', ''],
        [3, 'VIO-2024-003', '2024-02-01', 'وزارة الموارد البشرية', 'عمالية',
         'عدم توثيق عقود العمل', 'المادة 51 - نظام العمل', 10000, 'قائمة', ''],
        [4, 'VIO-2024-004', '2024-02-10', 'الدفاع المدني', 'سلامة',
         'انتهاء صلاحية طفايات الحريق', 'لائحة السلامة - البند 8', 15000, 'معالجة', '2024-02-20'],
        [5, 'VIO-2024-005', '2024-03-01', 'البلدية', 'بيئية',
         'عدم وجود عقد نظافة ساري', 'لائحة النظافة - المادة 5', 5000, 'قائمة', ''],
    ]

    for row_idx, violation in enumerate(violations, 4):
        for col_idx, value in enumerate(violation, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN
            cell.font = NORMAL_FONT

    # Data Validation للحالة
    status_dv = DataValidation(type="list", formula1='"قائمة,معالجة,مغلقة,متنازع عليها"', allow_blank=True)
    status_dv.error = "اختر من القائمة"
    status_dv.prompt = "اختر حالة المخالفة"
    ws1.add_data_validation(status_dv)
    status_dv.add('I4:I100')

    # Data Validation للجهة الرقابية
    authority_dv = DataValidation(type="list",
        formula1='"هيئة الغذاء والدواء,وزارة التجارة,وزارة الموارد البشرية,الدفاع المدني,البلدية,هيئة الزكاة والضريبة,الجهة البيئية"',
        allow_blank=True)
    ws1.add_data_validation(authority_dv)
    authority_dv.add('D4:D100')

    # Conditional Formatting للحالة
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws1.conditional_formatting.add('I4:I100',
        FormulaRule(formula=['$I4="قائمة"'], fill=red_fill))
    ws1.conditional_formatting.add('I4:I100',
        FormulaRule(formula=['$I4="معالجة"'], fill=yellow_fill))
    ws1.conditional_formatting.add('I4:I100',
        FormulaRule(formula=['$I4="مغلقة"'], fill=green_fill))

    set_column_widths(ws1, [5, 15, 12, 20, 15, 35, 25, 15, 12, 15])

    # ورقة 2: ملخص المخالفات
    ws2 = wb.create_sheet("ملخص المخالفات")
    ws2.sheet_view.rightToLeft = True

    ws2['A1'] = "ملخص المخالفات حسب الجهة"
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells('A1:D1')

    summary_headers = ['الجهة الرقابية', 'عدد المخالفات', 'إجمالي الغرامات', 'المخالفات القائمة']
    for col, header in enumerate(summary_headers, 1):
        ws2.cell(row=3, column=col, value=header)
    style_header_row(ws2, 3, len(summary_headers))

    # معادلات للحساب
    ws2['A4'] = 'هيئة الغذاء والدواء'
    ws2['B4'] = '=COUNTIF(\'سجل المخالفات\'!D:D,A4)'
    ws2['C4'] = '=SUMIF(\'سجل المخالفات\'!D:D,A4,\'سجل المخالفات\'!H:H)'
    ws2['D4'] = '=COUNTIFS(\'سجل المخالفات\'!D:D,A4,\'سجل المخالفات\'!I:I,"قائمة")'

    ws2['A5'] = 'وزارة الموارد البشرية'
    ws2['B5'] = '=COUNTIF(\'سجل المخالفات\'!D:D,A5)'
    ws2['C5'] = '=SUMIF(\'سجل المخالفات\'!D:D,A5,\'سجل المخالفات\'!H:H)'
    ws2['D5'] = '=COUNTIFS(\'سجل المخالفات\'!D:D,A5,\'سجل المخالفات\'!I:I,"قائمة")'

    ws2['A6'] = 'الدفاع المدني'
    ws2['B6'] = '=COUNTIF(\'سجل المخالفات\'!D:D,A6)'
    ws2['C6'] = '=SUMIF(\'سجل المخالفات\'!D:D,A6,\'سجل المخالفات\'!H:H)'
    ws2['D6'] = '=COUNTIFS(\'سجل المخالفات\'!D:D,A6,\'سجل المخالفات\'!I:I,"قائمة")'

    ws2['A7'] = 'البلدية'
    ws2['B7'] = '=COUNTIF(\'سجل المخالفات\'!D:D,A7)'
    ws2['C7'] = '=SUMIF(\'سجل المخالفات\'!D:D,A7,\'سجل المخالفات\'!H:H)'
    ws2['D7'] = '=COUNTIFS(\'سجل المخالفات\'!D:D,A7,\'سجل المخالفات\'!I:I,"قائمة")'

    # صف الإجمالي
    ws2['A9'] = 'الإجمالي'
    ws2['A9'].font = Font(bold=True)
    ws2['B9'] = '=SUM(B4:B8)'
    ws2['C9'] = '=SUM(C4:C8)'
    ws2['D9'] = '=SUM(D4:D8)'

    for row in range(4, 10):
        for col in range(1, 5):
            ws2.cell(row=row, column=col).border = THIN_BORDER
            ws2.cell(row=row, column=col).alignment = CENTER_ALIGN

    set_column_widths(ws2, [25, 15, 18, 18])

    # ورقة 3: العقوبات النظامية
    ws3 = wb.create_sheet("العقوبات النظامية")
    ws3.sheet_view.rightToLeft = True

    ws3['A1'] = "جدول العقوبات والغرامات حسب الأنظمة السعودية"
    ws3['A1'].font = TITLE_FONT
    ws3.merge_cells('A1:F1')

    penalty_headers = ['النظام', 'المخالفة', 'العقوبة - المرة الأولى', 'العقوبة - التكرار', 'الإغلاق', 'المرجع']
    for col, header in enumerate(penalty_headers, 1):
        ws3.cell(row=3, column=col, value=header)
    style_header_row(ws3, 3, len(penalty_headers))

    penalties = [
        ['نظام الغذاء', 'منتجات منتهية الصلاحية', '100,000 ريال', '500,000 ريال + إغلاق', 'نعم', 'المادة 23'],
        ['نظام الغذاء', 'عدم وجود ترخيص صحي', '50,000 ريال', '200,000 ريال', 'نعم', 'المادة 8'],
        ['نظام الغذاء', 'عدم تطبيق HACCP', '50,000 ريال', '150,000 ريال', 'نعم', 'المادة 15'],
        ['نظام العمل', 'عدم توثيق العقود', '10,000 ريال/عامل', '20,000 ريال/عامل', 'لا', 'المادة 51'],
        ['نظام العمل', 'تأخير الرواتب', '10,000 ريال', 'إيقاف الخدمات', 'لا', 'المادة 90'],
        ['نظام الشركات', 'عدم تجديد السجل', '10,000 ريال', '50,000 ريال', 'نعم', 'المادة 214'],
        ['الدفاع المدني', 'عدم توفر وسائل السلامة', '20,000 ريال', '50,000 ريال + إغلاق', 'نعم', 'لائحة السلامة'],
    ]

    for row_idx, penalty in enumerate(penalties, 4):
        for col_idx, value in enumerate(penalty, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN
            cell.font = NORMAL_FONT

    set_column_widths(ws3, [15, 25, 20, 25, 10, 15])

    # حفظ الملف
    wb.save('01_current_status/violation_register.xlsx')
    print("✓ تم إنشاء: violation_register.xlsx")

def create_suspended_licenses():
    """إنشاء سجل التراخيص الموقوفة"""
    wb = Workbook()

    # ورقة 1: التراخيص الموقوفة
    ws1 = wb.active
    ws1.title = "التراخيص"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:K1')
    ws1['A1'] = "سجل التراخيص والتصاريح"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER_ALIGN

    headers = ['م', 'نوع الترخيص', 'رقم الترخيص', 'الجهة المصدرة', 'تاريخ الإصدار',
               'تاريخ الانتهاء', 'أيام للانتهاء', 'الحالة', 'متطلبات التجديد', 'الرسوم', 'ملاحظات']

    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    licenses = [
        [1, 'السجل التجاري', '1010XXXXXX', 'وزارة التجارة', '2022-01-01', '2024-01-01',
         '=F4-TODAY()', 'منتهي', 'تجديد سنوي', 1200, 'يجب التجديد فوراً'],
        [2, 'الرخصة الصناعية', 'IND-XXXXX', 'وزارة الصناعة', '2021-06-01', '2024-06-01',
         '=F5-TODAY()', 'منتهي', 'شهادة بيئية + سلامة', 5000, 'مرتبط بالسجل التجاري'],
        [3, 'رخصة هيئة الغذاء والدواء', 'SFDA-XXXXX', 'هيئة الغذاء والدواء', '2022-03-01', '2024-03-01',
         '=F6-TODAY()', 'موقوف', 'تطبيق HACCP + فحص', 10000, 'سبب الإيقاف: مخالفات'],
        [4, 'شهادة الدفاع المدني', 'CD-XXXXX', 'الدفاع المدني', '2023-01-01', '2024-01-01',
         '=F7-TODAY()', 'منتهي', 'معاينة + صيانة معدات', 2000, 'فحص سنوي'],
        [5, 'رخصة البلدية', 'MUN-XXXXX', 'الأمانة/البلدية', '2023-03-01', '2024-03-01',
         '=F8-TODAY()', 'منتهي', 'سجل تجاري ساري + رسوم', 3000, ''],
        [6, 'شهادة الزكاة والدخل', 'ZATCA-XXXXX', 'هيئة الزكاة والضريبة', '2023-04-01', '2024-04-01',
         '=F9-TODAY()', 'ساري', 'إقرار ضريبي', 0, 'يجب التجديد قبل 3 أشهر'],
        [7, 'اشتراك التأمينات', 'GOSI-XXXXX', 'التأمينات الاجتماعية', '2023-01-01', '2024-12-31',
         '=F10-TODAY()', 'ساري', 'سداد الاشتراكات', 0, 'شهري'],
    ]

    for row_idx, lic in enumerate(licenses, 4):
        for col_idx, value in enumerate(lic, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN
            cell.font = NORMAL_FONT

    # Data Validation للحالة
    status_dv = DataValidation(type="list", formula1='"ساري,منتهي,موقوف,قيد التجديد"', allow_blank=True)
    ws1.add_data_validation(status_dv)
    status_dv.add('H4:H100')

    # Conditional Formatting
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws1.conditional_formatting.add('H4:H100',
        FormulaRule(formula=['OR($H4="منتهي",$H4="موقوف")'], fill=red_fill))
    ws1.conditional_formatting.add('H4:H100',
        FormulaRule(formula=['$H4="ساري"'], fill=green_fill))
    ws1.conditional_formatting.add('H4:H100',
        FormulaRule(formula=['$H4="قيد التجديد"'], fill=yellow_fill))

    # تنسيق عمود أيام للانتهاء
    ws1.conditional_formatting.add('G4:G100',
        FormulaRule(formula=['$G4<0'], fill=red_fill))
    ws1.conditional_formatting.add('G4:G100',
        FormulaRule(formula=['AND($G4>=0,$G4<=30)'], fill=yellow_fill))

    set_column_widths(ws1, [5, 22, 15, 20, 12, 12, 12, 12, 25, 10, 25])

    # ورقة 2: خطة التجديد
    ws2 = wb.create_sheet("خطة التجديد")
    ws2.sheet_view.rightToLeft = True

    ws2['A1'] = "خطة تجديد التراخيص - الأولوية حسب الأهمية"
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells('A1:G1')

    renewal_headers = ['الأولوية', 'الترخيص', 'الإجراء المطلوب', 'المدة المتوقعة', 'التكلفة', 'المسؤول', 'الموعد النهائي']
    for col, header in enumerate(renewal_headers, 1):
        ws2.cell(row=3, column=col, value=header)
    style_header_row(ws2, 3, len(renewal_headers))

    renewal_plan = [
        ['عاجل', 'السجل التجاري', 'تجديد فوري عبر منصة وزارة التجارة', '1-3 أيام', 1200, 'المدير المالي', ''],
        ['عاجل', 'رخصة هيئة الغذاء والدواء', 'معالجة المخالفات + طلب رفع الإيقاف', '30-60 يوم', 15000, 'مدير الجودة', ''],
        ['عاجل', 'شهادة الدفاع المدني', 'صيانة المعدات + طلب معاينة', '7-14 يوم', 5000, 'مدير السلامة', ''],
        ['مهم', 'الرخصة الصناعية', 'تجديد بعد السجل التجاري', '7-14 يوم', 5000, 'المدير المالي', ''],
        ['مهم', 'رخصة البلدية', 'تجديد بعد استكمال المتطلبات', '3-7 أيام', 3000, 'المدير المالي', ''],
    ]

    for row_idx, item in enumerate(renewal_plan, 4):
        for col_idx, value in enumerate(item, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN
            cell.font = NORMAL_FONT

    # Conditional Formatting للأولوية
    ws2.conditional_formatting.add('A4:A100',
        FormulaRule(formula=['$A4="عاجل"'], fill=red_fill))
    ws2.conditional_formatting.add('A4:A100',
        FormulaRule(formula=['$A4="مهم"'], fill=yellow_fill))

    set_column_widths(ws2, [10, 25, 35, 15, 12, 15, 15])

    wb.save('01_current_status/suspended_licenses.xlsx')
    print("✓ تم إنشاء: suspended_licenses.xlsx")

def create_gap_analysis():
    """إنشاء تحليل الفجوة الشامل"""
    wb = Workbook()

    # ورقة 1: تحليل الفجوة
    ws1 = wb.active
    ws1.title = "تحليل الفجوة"
    ws1.sheet_view.rightToLeft = True

    ws1.merge_cells('A1:I1')
    ws1['A1'] = "تحليل الفجوة الشامل - الوضع الراهن vs المطلوب"
    ws1['A1'].font = TITLE_FONT
    ws1['A1'].alignment = CENTER_ALIGN

    headers = ['م', 'المجال', 'المتطلب', 'الوضع الراهن', 'الوضع المطلوب',
               'الفجوة %', 'الأولوية', 'الإجراء التصحيحي', 'المدة']

    for col, header in enumerate(headers, 1):
        ws1.cell(row=3, column=col, value=header)
    style_header_row(ws1, 3, len(headers))

    gap_items = [
        [1, 'التراخيص', 'السجل التجاري ساري', 'منتهي', 'ساري', 100, 'حرجة', 'تجديد فوري', '3 أيام'],
        [2, 'التراخيص', 'رخصة SFDA سارية', 'موقوف', 'ساري', 100, 'حرجة', 'معالجة المخالفات', '60 يوم'],
        [3, 'سلامة الغذاء', 'نظام HACCP موثق', 'غير موجود', 'مطبق ومعتمد', 100, 'حرجة', 'تصميم وتطبيق النظام', '90 يوم'],
        [4, 'سلامة الغذاء', 'تحليل المخاطر', 'جزئي', 'مكتمل', 60, 'عالية', 'استكمال التحليل', '30 يوم'],
        [5, 'الموارد البشرية', 'عقود موثقة', '70%', '100%', 30, 'عالية', 'توثيق العقود المتبقية', '14 يوم'],
        [6, 'الموارد البشرية', 'تدريب سلامة الغذاء', '30%', '100%', 70, 'عالية', 'برنامج تدريب شامل', '45 يوم'],
        [7, 'البنية التحتية', 'معدات سلامة', '60%', '100%', 40, 'متوسطة', 'شراء وصيانة', '30 يوم'],
        [8, 'التوثيق', 'نظام ضبط الوثائق', 'غير موجود', 'مطبق', 100, 'عالية', 'تصميم النظام', '30 يوم'],
        [9, 'البيئة', 'تصريح بيئي', 'منتهي', 'ساري', 100, 'عالية', 'تجديد التصريح', '45 يوم'],
        [10, 'الجودة', 'مختبر داخلي', 'غير موجود', 'معتمد', 100, 'متوسطة', 'إنشاء أو تعاقد', '90 يوم'],
    ]

    for row_idx, item in enumerate(gap_items, 4):
        for col_idx, value in enumerate(item, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN
            cell.font = NORMAL_FONT

    # Data Validation للأولوية
    priority_dv = DataValidation(type="list", formula1='"حرجة,عالية,متوسطة,منخفضة"', allow_blank=True)
    ws1.add_data_validation(priority_dv)
    priority_dv.add('G4:G100')

    # Conditional Formatting للأولوية
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    ws1.conditional_formatting.add('G4:G100',
        FormulaRule(formula=['$G4="حرجة"'], fill=red_fill))
    ws1.conditional_formatting.add('G4:G100',
        FormulaRule(formula=['$G4="عالية"'], fill=orange_fill))
    ws1.conditional_formatting.add('G4:G100',
        FormulaRule(formula=['$G4="متوسطة"'], fill=yellow_fill))

    # Data Bar للفجوة
    ws1.conditional_formatting.add('F4:F100',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=100,
                   color="FF6B6B"))

    set_column_widths(ws1, [5, 15, 22, 18, 18, 10, 12, 30, 12])

    # ورقة 2: ملخص الفجوات
    ws2 = wb.create_sheet("ملخص الفجوات")
    ws2.sheet_view.rightToLeft = True

    ws2['A1'] = "ملخص الفجوات حسب المجال"
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells('A1:D1')

    summary_headers = ['المجال', 'عدد الفجوات', 'الفجوات الحرجة', 'متوسط الفجوة %']
    for col, header in enumerate(summary_headers, 1):
        ws2.cell(row=3, column=col, value=header)
    style_header_row(ws2, 3, len(summary_headers))

    ws2['A4'] = 'التراخيص'
    ws2['B4'] = '=COUNTIF(\'تحليل الفجوة\'!B:B,A4)'
    ws2['C4'] = '=COUNTIFS(\'تحليل الفجوة\'!B:B,A4,\'تحليل الفجوة\'!G:G,"حرجة")'
    ws2['D4'] = '=AVERAGEIF(\'تحليل الفجوة\'!B:B,A4,\'تحليل الفجوة\'!F:F)'

    ws2['A5'] = 'سلامة الغذاء'
    ws2['B5'] = '=COUNTIF(\'تحليل الفجوة\'!B:B,A5)'
    ws2['C5'] = '=COUNTIFS(\'تحليل الفجوة\'!B:B,A5,\'تحليل الفجوة\'!G:G,"حرجة")'
    ws2['D5'] = '=AVERAGEIF(\'تحليل الفجوة\'!B:B,A5,\'تحليل الفجوة\'!F:F)'

    ws2['A6'] = 'الموارد البشرية'
    ws2['B6'] = '=COUNTIF(\'تحليل الفجوة\'!B:B,A6)'
    ws2['C6'] = '=COUNTIFS(\'تحليل الفجوة\'!B:B,A6,\'تحليل الفجوة\'!G:G,"حرجة")'
    ws2['D6'] = '=AVERAGEIF(\'تحليل الفجوة\'!B:B,A6,\'تحليل الفجوة\'!F:F)'

    for row in range(4, 10):
        for col in range(1, 5):
            ws2.cell(row=row, column=col).border = THIN_BORDER
            ws2.cell(row=row, column=col).alignment = CENTER_ALIGN

    set_column_widths(ws2, [20, 15, 15, 15])

    # ورقة 3: خطة سد الفجوات
    ws3 = wb.create_sheet("خطة سد الفجوات")
    ws3.sheet_view.rightToLeft = True

    ws3['A1'] = "خطة سد الفجوات - الجدول الزمني"
    ws3['A1'].font = TITLE_FONT
    ws3.merge_cells('A1:H1')

    plan_headers = ['المرحلة', 'المدة', 'الفجوات المستهدفة', 'الإجراءات الرئيسية', 'المسؤول', 'الميزانية', 'مؤشر النجاح', 'الحالة']
    for col, header in enumerate(plan_headers, 1):
        ws3.cell(row=3, column=col, value=header)
    style_header_row(ws3, 3, len(plan_headers))

    phases = [
        ['الأسبوع 1-2', '14 يوم', 'التراخيص الحرجة', 'تجديد السجل + رفع الإيقاف', 'المدير العام', 25000, 'تراخيص سارية', 'لم يبدأ'],
        ['الأسبوع 3-4', '14 يوم', 'الموارد البشرية', 'توثيق العقود + التدريب', 'مدير الموارد البشرية', 15000, 'امتثال 100%', 'لم يبدأ'],
        ['الشهر 2', '30 يوم', 'سلامة الغذاء', 'تصميم نظام HACCP', 'مدير الجودة', 50000, 'نظام موثق', 'لم يبدأ'],
        ['الشهر 3', '30 يوم', 'البنية التحتية', 'تحديث المعدات والمرافق', 'مدير العمليات', 100000, 'مطابقة 100%', 'لم يبدأ'],
    ]

    for row_idx, phase in enumerate(phases, 4):
        for col_idx, value in enumerate(phase, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN
            cell.font = NORMAL_FONT

    # Data Validation للحالة
    status_dv = DataValidation(type="list", formula1='"لم يبدأ,قيد التنفيذ,مكتمل,متأخر"', allow_blank=True)
    ws3.add_data_validation(status_dv)
    status_dv.add('H4:H100')

    set_column_widths(ws3, [15, 10, 20, 30, 20, 12, 18, 12])

    wb.save('01_current_status/gap_analysis.xlsx')
    print("✓ تم إنشاء: gap_analysis.xlsx")

if __name__ == "__main__":
    os.chdir('/home/user/chtgpt/food_factory_compliance')
    print("جاري إنشاء ملفات الحالة الراهنة...")
    create_violation_register()
    create_suspended_licenses()
    create_gap_analysis()
    print("\n✓ تم إنشاء جميع ملفات 01_current_status بنجاح!")
