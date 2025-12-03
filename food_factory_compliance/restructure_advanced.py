#!/usr/bin/env python3
"""
سكربت إعادة هيكلة ملفات الامتثال - النسخة المتقدمة
يتضمن الأنظمة السعودية والشهادات الدولية والمتطلبات التفصيلية
"""

import os
import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

# =============================================================================
# الإعدادات الأساسية
# =============================================================================

SOURCE_DIR = "."
TARGET_DIR = "../food_factory_compliance_v2"

# الألوان
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
MANDATORY_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

# =============================================================================
# قاعدة بيانات الأنظمة والتشريعات السعودية
# =============================================================================

SAUDI_REGULATIONS = {
    "SFDA": {
        "name": "الهيئة العامة للغذاء والدواء",
        "laws": [
            {"code": "SFDA.FD/GSO 9", "name": "بطاقات المواد الغذائية المعبأة", "mandatory": True},
            {"code": "SFDA.FD/GSO 150-1", "name": "فترات الصلاحية الإلزامية", "mandatory": True},
            {"code": "SFDA.FD/GSO 150-2", "name": "فترات الصلاحية الاختيارية", "mandatory": False},
            {"code": "SFDA.FD/GSO 2055-1", "name": "اشتراطات الأغذية الحلال", "mandatory": True},
            {"code": "SFDA.FD 2333", "name": "الإدعاءات التغذوية والصحية", "mandatory": True},
            {"code": "SFDA Food Hygiene", "name": "اشتراطات صحة الغذاء", "mandatory": True},
        ],
        "systems": [
            {"name": "FIRS", "desc": "نظام تسجيل المنشآت الغذائية", "mandatory": True},
            {"name": "شهادة التصدير", "desc": "شهادة تصدير المنتجات الغذائية", "mandatory": True},
        ]
    },
    "SASO": {
        "name": "الهيئة السعودية للمواصفات والمقاييس والجودة",
        "laws": [
            {"code": "SASO GSO", "name": "المواصفات القياسية الخليجية", "mandatory": True},
            {"code": "SABER", "name": "منصة سابر لشهادات المطابقة", "mandatory": True},
            {"code": "SALEEM", "name": "برنامج سليم لسلامة المنتجات", "mandatory": True},
        ]
    },
    "MOHR": {
        "name": "وزارة الموارد البشرية والتنمية الاجتماعية",
        "laws": [
            {"code": "م/51", "name": "نظام العمل", "mandatory": True},
            {"code": "م/44", "name": "تعديلات نظام العمل 1446هـ", "mandatory": True},
            {"code": "نطاقات", "name": "برنامج نطاقات للسعودة", "mandatory": True},
            {"code": "حماية الأجور", "name": "نظام حماية الأجور", "mandatory": True},
        ],
        "articles": [
            {"num": "122", "topic": "ساعات العمل", "limit": "8 ساعات يومياً / 48 أسبوعياً"},
            {"num": "134", "topic": "السلامة المهنية", "limit": "توفير معدات الوقاية"},
            {"num": "137", "topic": "الفحص الطبي", "limit": "فحص دوري للعمال"},
            {"num": "150", "topic": "الإجازات السنوية", "limit": "21-30 يوم"},
        ],
        "courts": [
            {"stage": "التسوية الودية", "body": "مكتب العمل", "duration": "21 يوم", "mandatory": True},
            {"stage": "المحكمة العمالية", "body": "المحكمة العمالية", "duration": "متغير", "mandatory": False},
            {"stage": "الاستئناف", "body": "محكمة الاستئناف العمالية", "duration": "30 يوم", "mandatory": False},
        ]
    },
    "ENV": {
        "name": "وزارة البيئة والمياه والزراعة",
        "laws": [
            {"code": "نظام البيئة", "name": "نظام البيئة السعودي", "mandatory": True},
            {"code": "التصريح البيئي", "name": "تصريح بيئي للمنشآت", "mandatory": True},
            {"code": "NCEC", "name": "المركز الوطني للرقابة البيئية", "mandatory": True},
        ],
        "categories": [
            {"cat": "الفئة الأولى", "desc": "مشروعات ذات تأثير محدود", "examples": "ورش صغيرة"},
            {"cat": "الفئة الثانية", "desc": "مشروعات متوسطة التأثير", "examples": "مصانع الأغذية"},
            {"cat": "الفئة الثالثة", "desc": "مشروعات عالية التأثير", "examples": "مصانع كبيرة"},
        ]
    },
    "CIVIL_DEFENSE": {
        "name": "الدفاع المدني",
        "laws": [
            {"code": "م/10", "name": "نظام الدفاع المدني", "mandatory": True},
            {"code": "سلامة", "name": "بوابة سلامة الإلكترونية", "mandatory": True},
            {"code": "SBC", "name": "كود البناء السعودي", "mandatory": True},
        ],
        "requirements": [
            {"item": "أنظمة إنذار الحريق", "mandatory": True},
            {"item": "أنظمة إطفاء الحريق", "mandatory": True},
            {"item": "مخارج الطوارئ", "mandatory": True},
            {"item": "خطط الإخلاء", "mandatory": True},
            {"item": "طفايات الحريق", "mandatory": True},
            {"item": "التهوية والإضاءة", "mandatory": True},
        ]
    },
    "MODON": {
        "name": "الهيئة السعودية للمدن الصناعية (مدن)",
        "laws": [
            {"code": "ترخيص المصنع", "name": "رخصة تشغيل المصنع", "mandatory": True},
            {"code": "البيئية الصناعية", "name": "الاشتراطات البيئية الصناعية", "mandatory": True},
        ]
    }
}

# =============================================================================
# قاعدة بيانات الشهادات الدولية
# =============================================================================

INTERNATIONAL_CERTIFICATIONS = {
    "GFSI_RECOGNIZED": {
        "name": "شهادات معتمدة من GFSI",
        "certs": [
            {
                "code": "FSSC 22000",
                "name": "شهادة نظام سلامة الغذاء",
                "components": ["ISO 22000", "ISO/TS 22002-1", "HACCP", "متطلبات إضافية"],
                "validity": "3 سنوات",
                "markets": ["أوروبا", "أمريكا", "آسيا", "الخليج"],
                "priority": 1
            },
            {
                "code": "BRC",
                "name": "المعيار العالمي لسلامة الغذاء",
                "components": ["التزام الإدارة العليا", "HACCP", "نظام الجودة", "بيئة المصنع"],
                "validity": "سنة واحدة",
                "markets": ["أوروبا", "المملكة المتحدة", "سلاسل التجزئة العالمية"],
                "priority": 2
            },
            {
                "code": "IFS",
                "name": "المعيار الدولي للغذاء",
                "components": ["HACCP", "إدارة الجودة", "سلامة المنتج"],
                "validity": "سنة واحدة",
                "markets": ["ألمانيا", "فرنسا", "إيطاليا", "أوروبا"],
                "priority": 3
            },
            {
                "code": "SQF",
                "name": "الغذاء الآمن والجودة",
                "components": ["نظام إدارة الجودة", "HACCP", "ممارسات التصنيع الجيد"],
                "validity": "سنة واحدة",
                "markets": ["أمريكا الشمالية", "أستراليا"],
                "priority": 4
            },
        ]
    },
    "ISO_STANDARDS": {
        "name": "معايير الأيزو",
        "certs": [
            {
                "code": "ISO 22000:2018",
                "name": "نظام إدارة سلامة الغذاء",
                "clauses": [
                    "4. سياق المنظمة",
                    "5. القيادة",
                    "6. التخطيط",
                    "7. الدعم",
                    "8. التشغيل",
                    "9. تقييم الأداء",
                    "10. التحسين"
                ],
                "mandatory": False
            },
            {
                "code": "ISO 9001:2015",
                "name": "نظام إدارة الجودة",
                "mandatory": False
            },
            {
                "code": "ISO 14001:2015",
                "name": "نظام الإدارة البيئية",
                "mandatory": False
            },
            {
                "code": "ISO 45001:2018",
                "name": "نظام إدارة السلامة والصحة المهنية",
                "mandatory": False
            },
        ]
    },
    "HACCP": {
        "name": "نظام تحليل المخاطر ونقاط التحكم الحرجة",
        "principles": [
            "1. تحليل المخاطر",
            "2. تحديد نقاط التحكم الحرجة (CCPs)",
            "3. وضع الحدود الحرجة",
            "4. إجراءات المراقبة",
            "5. الإجراءات التصحيحية",
            "6. إجراءات التحقق",
            "7. التوثيق وحفظ السجلات"
        ],
        "mandatory": True
    },
    "PRPs": {
        "name": "البرامج الشرطية المسبقة",
        "programs": [
            "تصميم وصيانة المباني",
            "تصميم المعدات",
            "إدارة المواد المشتراة",
            "التنظيف والتطهير",
            "مكافحة الآفات",
            "النظافة الشخصية",
            "إدارة النفايات",
            "إدارة المياه",
            "التحكم في درجات الحرارة",
            "جودة الهواء والتهوية"
        ]
    }
}

# =============================================================================
# متطلبات التأهيل والتدريب
# =============================================================================

TRAINING_REQUIREMENTS = {
    "mandatory": [
        {"course": "HACCP المستوى الأول", "hours": 16, "validity": "3 سنوات", "target": "جميع العاملين"},
        {"course": "HACCP المستوى الثاني", "hours": 24, "validity": "3 سنوات", "target": "المشرفين"},
        {"course": "سلامة الغذاء الأساسية", "hours": 8, "validity": "سنتان", "target": "العمال"},
        {"course": "النظافة الشخصية", "hours": 4, "validity": "سنوياً", "target": "الجميع"},
        {"course": "مكافحة الآفات", "hours": 8, "validity": "سنتان", "target": "المختصين"},
        {"course": "السلامة والصحة المهنية", "hours": 16, "validity": "سنتان", "target": "الجميع"},
        {"course": "الإسعافات الأولية", "hours": 16, "validity": "سنتان", "target": "10% من العمالة"},
        {"course": "إطفاء الحريق", "hours": 8, "validity": "سنوياً", "target": "فريق الطوارئ"},
    ],
    "recommended": [
        {"course": "المدقق الداخلي ISO 22000", "hours": 24, "target": "فريق الجودة"},
        {"course": "إدارة المخاطر", "hours": 16, "target": "الإدارة"},
        {"course": "إدارة الأزمات والاستدعاء", "hours": 8, "target": "الإدارة العليا"},
        {"course": "التدقيق على الموردين", "hours": 16, "target": "المشتريات"},
    ]
}

# =============================================================================
# هيكل الملفات الجديد
# =============================================================================

FOLDER_MAPPING = {
    "01_current_status": "01_الحالة_الراهنة.xlsx",
    "02_regulatory_framework": "02_الإطار_التنظيمي.xlsx",
    "03_compliance_roadmap": "03_خارطة_الامتثال.xlsx",
    "04_certifications": "04_الشهادات_الدولية.xlsx",
    "05_export_readiness": "05_جاهزية_التصدير.xlsx",
    "06_risk_management": "06_إدارة_المخاطر.xlsx",
    "07_kpis_dashboard": "07_لوحة_المؤشرات.xlsx",
    "08_documentation": "08_التوثيق.xlsx",
    "09_financial_planning": "09_التخطيط_المالي.xlsx",
    "10_checklists": "10_قوائم_الفحص.xlsx",
    "11_master_files": "11_الملفات_الرئيسية.xlsx",
}

NEW_FILES = {
    "09_الموارد_البشرية_والتأهيل.xlsx": [
        "الهيكل_الوظيفي",
        "مصفوفة_الصلاحيات",
        "خطة_التدريب_السنوية",
        "سجل_الشهادات_المهنية",
        "تقييم_الكفاءات",
        "متطلبات_نظام_العمل"
    ],
    "10_إدارة_الموردين.xlsx": [
        "سجل_الموردين",
        "معايير_التأهيل",
        "جدول_التدقيق",
        "اتفاقيات_الجودة",
        "القائمة_السوداء",
        "تتبع_المواد_الخام"
    ],
    "11_إدارة_الأزمات_والسحب.xlsx": [
        "فريق_الأزمات",
        "خطة_الاستدعاء",
        "سيناريوهات_الطوارئ",
        "نماذج_الإبلاغ",
        "سجل_الحوادث",
        "التواصل_الأزمات"
    ]
}

DASHBOARD_EXTRA_SHEETS = [
    "الربط_المركزي",
    "التنبيهات",
    "مؤشر_الجاهزية_الكلي",
    "خارطة_السنة",
    "مواعيد_التجديد"
]

GOVERNANCE_EXTRA_SHEETS = [
    "جدول_التدقيق_السنوي",
    "تقارير_عدم_المطابقة",
    "الإجراءات_التصحيحية",
    "سجل_المراجعات"
]

# =============================================================================
# الدوال المساعدة
# =============================================================================

def style_header(ws, row, cols, text, merge=True):
    """تنسيق عنوان رئيسي"""
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='right', vertical='center')
    if merge and cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)

def style_subheader(ws, row, col, text):
    """تنسيق عنوان فرعي"""
    cell = ws.cell(row=row, column=col)
    cell.value = text
    cell.font = Font(color="FFFFFF", bold=True)
    cell.fill = SUBHEADER_FILL

def add_data_row(ws, row, data, mandatory=None):
    """إضافة صف بيانات"""
    for col, value in enumerate(data, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        if mandatory is not None:
            cell.fill = MANDATORY_FILL if mandatory else OPTIONAL_FILL

def set_column_widths(ws, widths):
    """ضبط عرض الأعمدة"""
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

def create_target_directory():
    """إنشاء المجلد الهدف"""
    if os.path.exists(TARGET_DIR):
        backup_name = f"{TARGET_DIR}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shutil.move(TARGET_DIR, backup_name)
        print(f"✓ نسخة احتياطية: {backup_name}")
    os.makedirs(TARGET_DIR)
    print(f"✓ تم إنشاء: {TARGET_DIR}")

# =============================================================================
# إنشاء ملف الأنظمة والتشريعات
# =============================================================================

def create_regulations_reference():
    """إنشاء ملف مرجعي للأنظمة والتشريعات"""
    filepath = os.path.join(TARGET_DIR, "00_مرجع_الأنظمة_والتشريعات.xlsx")
    wb = Workbook()

    # === ورقة الأنظمة السعودية ===
    ws1 = wb.active
    ws1.title = "الأنظمة_السعودية"
    ws1.sheet_view.rightToLeft = True

    row = 1
    style_header(ws1, row, 5, "📋 الأنظمة والتشريعات السعودية لمصانع الأغذية")
    row += 2

    for agency_code, agency_data in SAUDI_REGULATIONS.items():
        style_header(ws1, row, 5, f"🏛️ {agency_data['name']} ({agency_code})")
        row += 1

        # العناوين الفرعية
        headers = ["الرمز", "الاسم", "إلزامي", "الحالة", "ملاحظات"]
        for col, header in enumerate(headers, 1):
            style_subheader(ws1, row, col, header)
        row += 1

        # البيانات
        if "laws" in agency_data:
            for law in agency_data["laws"]:
                mandatory = law.get("mandatory", True)
                add_data_row(ws1, row, [
                    law["code"],
                    law["name"],
                    "نعم" if mandatory else "لا",
                    "⬜ غير مطبق",
                    ""
                ], mandatory)
                row += 1

        if "requirements" in agency_data:
            for req in agency_data["requirements"]:
                add_data_row(ws1, row, [
                    "-",
                    req["item"],
                    "نعم" if req["mandatory"] else "لا",
                    "⬜ غير مطبق",
                    ""
                ], req["mandatory"])
                row += 1

        row += 1

    set_column_widths(ws1, [20, 40, 10, 15, 30])

    # === ورقة الشهادات الدولية ===
    ws2 = wb.create_sheet("الشهادات_الدولية")
    ws2.sheet_view.rightToLeft = True

    row = 1
    style_header(ws2, row, 6, "🌍 الشهادات الدولية لسلامة الغذاء")
    row += 2

    # شهادات GFSI
    style_header(ws2, row, 6, "شهادات معتمدة من GFSI (المبادرة العالمية لسلامة الغذاء)")
    row += 1

    headers = ["الشهادة", "الاسم", "المكونات", "الصلاحية", "الأسواق المستهدفة", "الأولوية"]
    for col, header in enumerate(headers, 1):
        style_subheader(ws2, row, col, header)
    row += 1

    for cert in INTERNATIONAL_CERTIFICATIONS["GFSI_RECOGNIZED"]["certs"]:
        add_data_row(ws2, row, [
            cert["code"],
            cert["name"],
            " | ".join(cert["components"]),
            cert["validity"],
            " | ".join(cert["markets"]),
            f"#{cert['priority']}"
        ])
        row += 1

    row += 1

    # معايير ISO
    style_header(ws2, row, 6, "معايير الأيزو")
    row += 1

    for col, header in enumerate(["الرمز", "الاسم", "إلزامي", "الحالة", "", ""], 1):
        if header:
            style_subheader(ws2, row, col, header)
    row += 1

    for cert in INTERNATIONAL_CERTIFICATIONS["ISO_STANDARDS"]["certs"]:
        add_data_row(ws2, row, [
            cert["code"],
            cert["name"],
            "لا" if not cert.get("mandatory", False) else "نعم",
            "⬜ غير حاصل",
            "",
            ""
        ])
        row += 1

    set_column_widths(ws2, [20, 35, 40, 15, 30, 10])

    # === ورقة HACCP ===
    ws3 = wb.create_sheet("مبادئ_HACCP")
    ws3.sheet_view.rightToLeft = True

    row = 1
    style_header(ws3, row, 4, "🔬 نظام تحليل المخاطر ونقاط التحكم الحرجة (HACCP)")
    row += 2

    style_header(ws3, row, 4, "المبادئ السبعة لـ HACCP")
    row += 1

    headers = ["المبدأ", "الوصف", "المسؤول", "الحالة"]
    for col, header in enumerate(headers, 1):
        style_subheader(ws3, row, col, header)
    row += 1

    haccp_details = [
        ("1. تحليل المخاطر", "تحديد المخاطر البيولوجية والكيميائية والفيزيائية"),
        ("2. تحديد CCPs", "تحديد نقاط التحكم الحرجة في العملية"),
        ("3. الحدود الحرجة", "وضع حدود قابلة للقياس لكل CCP"),
        ("4. نظام المراقبة", "إجراءات مراقبة مستمرة للـ CCPs"),
        ("5. الإجراءات التصحيحية", "خطوات عند تجاوز الحدود الحرجة"),
        ("6. إجراءات التحقق", "التأكد من فعالية النظام"),
        ("7. التوثيق", "حفظ جميع السجلات والوثائق"),
    ]

    for principle, desc in haccp_details:
        add_data_row(ws3, row, [principle, desc, "", "⬜"])
        row += 1

    row += 2

    # PRPs
    style_header(ws3, row, 4, "البرامج الشرطية المسبقة (PRPs)")
    row += 1

    for col, header in enumerate(["البرنامج", "الوصف", "المسؤول", "الحالة"], 1):
        style_subheader(ws3, row, col, header)
    row += 1

    for prp in INTERNATIONAL_CERTIFICATIONS["PRPs"]["programs"]:
        add_data_row(ws3, row, [prp, "", "", "⬜"])
        row += 1

    set_column_widths(ws3, [25, 45, 20, 10])

    # === ورقة متطلبات التدريب ===
    ws4 = wb.create_sheet("متطلبات_التدريب")
    ws4.sheet_view.rightToLeft = True

    row = 1
    style_header(ws4, row, 5, "📚 متطلبات التدريب والتأهيل")
    row += 2

    style_header(ws4, row, 5, "الدورات الإلزامية")
    row += 1

    headers = ["الدورة", "الساعات", "الصلاحية", "الفئة المستهدفة", "الحالة"]
    for col, header in enumerate(headers, 1):
        style_subheader(ws4, row, col, header)
    row += 1

    for course in TRAINING_REQUIREMENTS["mandatory"]:
        add_data_row(ws4, row, [
            course["course"],
            course["hours"],
            course["validity"],
            course["target"],
            "⬜"
        ], True)
        row += 1

    row += 1
    style_header(ws4, row, 5, "الدورات الموصى بها")
    row += 1

    for col, header in enumerate(headers, 1):
        style_subheader(ws4, row, col, header)
    row += 1

    for course in TRAINING_REQUIREMENTS["recommended"]:
        add_data_row(ws4, row, [
            course["course"],
            course["hours"],
            "-",
            course["target"],
            "⬜"
        ], False)
        row += 1

    set_column_widths(ws4, [30, 10, 15, 25, 10])

    # === ورقة المحاكم العمالية ===
    ws5 = wb.create_sheet("المحاكم_العمالية")
    ws5.sheet_view.rightToLeft = True

    row = 1
    style_header(ws5, row, 5, "⚖️ المحاكم العمالية (النظام الجديد 2023)")
    row += 2

    ws5.cell(row=row, column=1).value = "ملاحظة: تم إلغاء هيئات تسوية الخلافات العمالية وإنشاء المحاكم العمالية"
    ws5.cell(row=row, column=1).font = Font(italic=True, color="C00000", bold=True)
    row += 2

    headers = ["المرحلة", "الجهة", "المدة", "إلزامي", "ملاحظات"]
    for col, header in enumerate(headers, 1):
        style_subheader(ws5, row, col, header)
    row += 1

    for court in SAUDI_REGULATIONS["MOHR"]["courts"]:
        add_data_row(ws5, row, [
            court["stage"],
            court["body"],
            court["duration"],
            "نعم" if court["mandatory"] else "لا",
            ""
        ], court["mandatory"])
        row += 1

    row += 2
    ws5.cell(row=row, column=1).value = "📌 التقاضي الإلكتروني عبر منصة ناجز (najiz.sa)"
    row += 1
    ws5.cell(row=row, column=1).value = "📌 الأحكام ملزمة وقابلة للتنفيذ مباشرة عبر محكمة التنفيذ"

    set_column_widths(ws5, [20, 25, 15, 10, 30])

    # === ورقة الدفاع المدني ===
    ws6 = wb.create_sheet("اشتراطات_الدفاع_المدني")
    ws6.sheet_view.rightToLeft = True

    row = 1
    style_header(ws6, row, 4, "🚒 اشتراطات الدفاع المدني للمصانع")
    row += 2

    headers = ["المتطلب", "الوصف", "إلزامي", "الحالة"]
    for col, header in enumerate(headers, 1):
        style_subheader(ws6, row, col, header)
    row += 1

    for req in SAUDI_REGULATIONS["CIVIL_DEFENSE"]["requirements"]:
        add_data_row(ws6, row, [
            req["item"],
            "",
            "نعم" if req["mandatory"] else "لا",
            "⬜"
        ], req["mandatory"])
        row += 1

    row += 2
    ws6.cell(row=row, column=1).value = "📌 ملاحظة: يتم إصدار رخصة الدفاع المدني عبر بوابة سلامة الإلكترونية"
    ws6.cell(row=row, column=1).font = Font(italic=True, color="666666")

    set_column_widths(ws6, [25, 40, 10, 10])

    # حفظ الملف
    wb.save(filepath)
    wb.close()
    print(f"✓ ملف مرجعي: {filepath}")

# =============================================================================
# نسخ ودمج الملفات
# =============================================================================

def copy_existing_files():
    """نسخ الملفات الموجودة"""
    print("\n--- نسخ الملفات الموجودة ---")

    for folder, filename in FOLDER_MAPPING.items():
        source_path = os.path.join(folder, filename)
        target_path = os.path.join(TARGET_DIR, filename)

        if os.path.exists(source_path):
            shutil.copy2(source_path, target_path)
            print(f"✓ {filename}")
        else:
            print(f"✗ غير موجود: {source_path}")

def add_sheets_to_workbook(filename, sheet_names):
    """إضافة أوراق جديدة"""
    filepath = os.path.join(TARGET_DIR, filename)

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.sheet_view.rightToLeft = True
            ws["A1"] = sheet_name
            ws["A1"].font = Font(bold=True, size=14)

    wb.save(filepath)
    wb.close()
    print(f"✓ أضيفت أوراق إلى: {filename}")

def create_new_file(filename, sheet_names):
    """إنشاء ملف جديد"""
    filepath = os.path.join(TARGET_DIR, filename)

    wb = Workbook()

    for sheet_name in sheet_names:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.rightToLeft = True
        ws["A1"] = sheet_name
        ws["A1"].font = Font(bold=True, size=14)
        ws["A3"] = "# أضف البيانات هنا"

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)
    wb.close()
    print(f"✓ ملف جديد: {filename}")

def generate_summary():
    """إنشاء ملخص"""
    summary = []
    summary.append("=" * 60)
    summary.append("📊 ملخص الهيكل الجديد")
    summary.append("=" * 60)
    summary.append("")

    for filename in sorted(os.listdir(TARGET_DIR)):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(TARGET_DIR, filename)
            wb = load_workbook(filepath, read_only=True)
            sheets = wb.sheetnames
            wb.close()

            summary.append(f"📁 {filename} ({len(sheets)} ورقة)")
            for sheet in sheets[:5]:
                summary.append(f"   └── {sheet}")
            if len(sheets) > 5:
                summary.append(f"   └── ... و {len(sheets)-5} أوراق أخرى")
            summary.append("")

    return "\n".join(summary)

# =============================================================================
# البرنامج الرئيسي
# =============================================================================

def main():
    print("=" * 60)
    print("🏭 سكربت إعادة هيكلة ملفات الامتثال - النسخة المتقدمة")
    print("   يتضمن الأنظمة السعودية والشهادات الدولية")
    print("=" * 60)
    print()

    # 1. إنشاء المجلد الهدف
    create_target_directory()
    print()

    # 2. إنشاء ملف الأنظمة والتشريعات
    print("--- إنشاء ملف الأنظمة والتشريعات ---")
    create_regulations_reference()
    print()

    # 3. نسخ الملفات الموجودة
    copy_existing_files()
    print()

    # 4. إضافة أوراق للوحة القيادة
    print("--- تحسين لوحة القيادة ---")
    add_sheets_to_workbook("07_لوحة_المؤشرات.xlsx", DASHBOARD_EXTRA_SHEETS)
    print()

    # 5. إضافة أوراق للتوثيق
    print("--- تحسين التوثيق ---")
    add_sheets_to_workbook("08_التوثيق.xlsx", GOVERNANCE_EXTRA_SHEETS)
    print()

    # 6. إنشاء الملفات الجديدة
    print("--- إنشاء الملفات الجديدة ---")
    for filename, sheets in NEW_FILES.items():
        create_new_file(filename, sheets)
    print()

    # 7. الملخص
    print(generate_summary())

    print("=" * 60)
    print("✅ اكتملت إعادة الهيكلة بنجاح!")
    print(f"📂 المجلد الجديد: {TARGET_DIR}")
    print()
    print("📋 الإضافات الجديدة:")
    print("   • ملف مرجعي للأنظمة السعودية (SFDA, SASO, MOHR, البيئة, الدفاع المدني)")
    print("   • الشهادات الدولية (FSSC 22000, BRC, IFS, ISO)")
    print("   • مبادئ HACCP والبرامج الشرطية PRPs")
    print("   • متطلبات التدريب والتأهيل")
    print("   • المحاكم العمالية (النظام الجديد 2023)")
    print("   • اشتراطات الدفاع المدني")
    print("=" * 60)

if __name__ == "__main__":
    main()
