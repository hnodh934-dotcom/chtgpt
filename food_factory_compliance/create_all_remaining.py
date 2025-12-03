#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
إنشاء جميع الملفات المتبقية للمشروع
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, Reference
import os

# الإعدادات العامة
GREEN = "1B5E4A"
GOLD = "C9A227"
HEADER_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
HEADER_FONT = Font(name='Arial', size=12, bold=True, color="FFFFFF")
TITLE_FONT = Font(name='Arial', size=16, bold=True, color=GREEN)
NORMAL_FONT = Font(name='Arial', size=11)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
RTL_ALIGN = Alignment(horizontal='right', vertical='center', wrap_text=True)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_data(ws, start_row, data):
    for row_idx, row_data in enumerate(data, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = RTL_ALIGN if col_idx > 1 else CENTER_ALIGN
            cell.font = NORMAL_FONT

# ===================== 03_compliance_roadmap =====================
def create_roadmap():
    print("  جاري إنشاء خارطة الامتثال...")

    # المرحلة 1: العاجلة 30 يوم
    wb = Workbook()
    ws = wb.active
    ws.title = "المرحلة العاجلة"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "المرحلة العاجلة: رفع الإيقاف (30 يوم)"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')

    headers = ['م', 'المهمة', 'المسؤول', 'الأسبوع 1', 'الأسبوع 2', 'الأسبوع 3', 'الأسبوع 4', 'الحالة']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    tasks = [
        [1, 'تجديد السجل التجاري', 'المدير المالي', '●', '', '', '', ''],
        [2, 'سداد المخالفات المالية', 'المدير المالي', '●', '', '', '', ''],
        [3, 'تجهيز ملف رفع الإيقاف', 'مدير الجودة', '●', '●', '', '', ''],
        [4, 'معالجة مخالفات SFDA', 'مدير الجودة', '●', '●', '●', '', ''],
        [5, 'تجديد شهادة الدفاع المدني', 'مدير السلامة', '', '●', '●', '', ''],
        [6, 'تجديد رخصة البلدية', 'المدير المالي', '', '', '●', '●', ''],
        [7, 'تقديم طلب رفع الإيقاف', 'المدير العام', '', '', '', '●', ''],
        [8, 'متابعة الطلب والرد', 'المدير العام', '', '', '', '●', ''],
    ]
    add_data(ws, 4, tasks)

    status_dv = DataValidation(type="list", formula1='"لم يبدأ,قيد التنفيذ,مكتمل,متأخر"')
    ws.add_data_validation(status_dv)
    status_dv.add('H4:H50')

    set_widths(ws, [5, 30, 18, 12, 12, 12, 12, 12])
    wb.save('03_compliance_roadmap/phase1_emergency_30days.xlsx')
    print("    ✓ phase1_emergency_30days.xlsx")

    # المرحلة 2: التثبيت 90 يوم
    wb = Workbook()
    ws = wb.active
    ws.title = "مرحلة التثبيت"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "مرحلة التثبيت: التشغيل النظامي (90 يوم)"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    headers = ['م', 'المهمة', 'المسؤول', 'الشهر 1', 'الشهر 2', 'الشهر 3', 'الحالة']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    tasks = [
        [1, 'تصميم نظام HACCP', 'مدير الجودة', '●', '●', '', ''],
        [2, 'تدريب الموظفين على سلامة الغذاء', 'الموارد البشرية', '●', '●', '', ''],
        [3, 'إنشاء نظام ضبط الوثائق', 'مدير الجودة', '●', '', '', ''],
        [4, 'تطبيق GMP و GHP', 'مدير الإنتاج', '●', '●', '●', ''],
        [5, 'إنشاء نظام التتبع', 'مدير الجودة', '', '●', '●', ''],
        [6, 'التعاقد مع مختبر معتمد', 'مدير الجودة', '●', '', '', ''],
        [7, 'برنامج مكافحة الآفات', 'مدير المرافق', '●', '', '', ''],
        [8, 'المراجعة الداخلية الأولى', 'مدير الجودة', '', '', '●', ''],
    ]
    add_data(ws, 4, tasks)
    set_widths(ws, [5, 35, 18, 12, 12, 12, 12])
    wb.save('03_compliance_roadmap/phase2_stabilization_90days.xlsx')
    print("    ✓ phase2_stabilization_90days.xlsx")

    # المرحلة 3: التحسين 180 يوم
    wb = Workbook()
    ws = wb.active
    ws.title = "مرحلة التحسين"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "مرحلة التحسين: الجودة والكفاءة (180 يوم)"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')

    headers = ['م', 'المهمة', 'المسؤول', 'ش1-2', 'ش3-4', 'ش5-6', 'الميزانية', 'الحالة']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    tasks = [
        [1, 'تطبيق ISO 22000', 'مدير الجودة', '●', '●', '●', 80000, ''],
        [2, 'تحديث المعدات', 'مدير الإنتاج', '●', '●', '', 150000, ''],
        [3, 'إنشاء مختبر داخلي', 'مدير الجودة', '●', '●', '', 100000, ''],
        [4, 'برنامج تطوير الموظفين', 'الموارد البشرية', '●', '●', '●', 50000, ''],
        [5, 'نظام إدارة الشكاوى', 'خدمة العملاء', '●', '', '', 10000, ''],
        [6, 'تحسين كفاءة الطاقة', 'مدير المرافق', '', '●', '●', 75000, ''],
    ]
    add_data(ws, 4, tasks)
    set_widths(ws, [5, 30, 18, 10, 10, 10, 15, 12])
    wb.save('03_compliance_roadmap/phase3_optimization_180days.xlsx')
    print("    ✓ phase3_optimization_180days.xlsx")

    # المرحلة 4: التميز 365 يوم
    wb = Workbook()
    ws = wb.active
    ws.title = "مرحلة التميز"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "مرحلة التميز: الشهادات الدولية (سنة)"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    headers = ['م', 'الشهادة', 'الجهة المانحة', 'المدة', 'التكلفة', 'الأولوية', 'الحالة']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    tasks = [
        [1, 'ISO 22000:2018', 'SGS / TUV', '6 أشهر', 80000, 'عالية', ''],
        [2, 'HACCP معتمد', 'هيئة معتمدة', '3 أشهر', 30000, 'عالية', ''],
        [3, 'شهادة الحلال', 'MUI / JAKIM', '3 أشهر', 25000, 'عالية', ''],
        [4, 'ISO 9001:2015', 'SGS / TUV', '6 أشهر', 60000, 'متوسطة', ''],
        [5, 'BRC Global', 'BRC', '9 أشهر', 100000, 'للتصدير', ''],
        [6, 'FSSC 22000', 'FSSC', '9 أشهر', 120000, 'للتصدير', ''],
    ]
    add_data(ws, 4, tasks)
    set_widths(ws, [5, 22, 18, 12, 15, 12, 12])
    wb.save('03_compliance_roadmap/phase4_excellence_365days.xlsx')
    print("    ✓ phase4_excellence_365days.xlsx")

    # المرحلة 5: العالمية 24 شهر
    wb = Workbook()
    ws = wb.active
    ws.title = "مرحلة العالمية"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "مرحلة العالمية: التصدير والمعارض (24 شهر)"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    headers = ['م', 'الهدف', 'السوق المستهدف', 'المتطلبات', 'الجدول الزمني', 'الميزانية', 'الحالة']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    tasks = [
        [1, 'التصدير لدول الخليج', 'GCC', 'GSO + شهادة منشأ', 'شهر 12-15', 50000, ''],
        [2, 'التصدير للاتحاد الأوروبي', 'EU', 'CE + HACCP + EU requirements', 'شهر 18-24', 200000, ''],
        [3, 'التصدير لأمريكا', 'USA', 'FDA registration + FSMA', 'شهر 18-24', 250000, ''],
        [4, 'معرض جلفود دبي', 'دولي', 'تجهيز الجناح + المنتجات', 'فبراير سنوياً', 150000, ''],
        [5, 'معرض أنوجا ألمانيا', 'EU', 'تجهيز الجناح', 'أكتوبر سنوياً', 200000, ''],
        [6, 'التجارة الإلكترونية الدولية', 'عالمي', 'منصات + لوجستيات', 'شهر 15-18', 100000, ''],
    ]
    add_data(ws, 4, tasks)
    set_widths(ws, [5, 28, 18, 30, 15, 15, 12])
    wb.save('03_compliance_roadmap/phase5_global_24months.xlsx')
    print("    ✓ phase5_global_24months.xlsx")

# ===================== 04_certifications =====================
def create_certifications():
    print("  جاري إنشاء ملفات الشهادات...")

    certs = [
        ('iso_22000_roadmap.xlsx', 'ISO 22000 - نظام إدارة سلامة الغذاء', [
            [1, 'تشكيل فريق سلامة الغذاء', 'أسبوع 1-2', 'مدير الجودة', ''],
            [2, 'تحليل المخاطر الأولي', 'أسبوع 2-4', 'الفريق', ''],
            [3, 'تحديد PRPs', 'أسبوع 3-6', 'مدير الجودة', ''],
            [4, 'تحديد CCPs', 'أسبوع 5-8', 'الفريق', ''],
            [5, 'إعداد الوثائق', 'أسبوع 6-16', 'مدير الجودة', ''],
            [6, 'التطبيق والتدريب', 'أسبوع 12-20', 'الفريق', ''],
            [7, 'المراجعة الداخلية', 'أسبوع 18-22', 'المراجع الداخلي', ''],
            [8, 'مراجعة الإدارة', 'أسبوع 22-24', 'المدير العام', ''],
            [9, 'تدقيق الشهادة', 'أسبوع 24-26', 'جهة الاعتماد', ''],
        ]),
        ('haccp_implementation.xlsx', 'HACCP - تحليل المخاطر ونقاط التحكم الحرجة', [
            [1, 'تجميع فريق HACCP', 'أسبوع 1', 'المدير العام', ''],
            [2, 'وصف المنتج والاستخدام المقصود', 'أسبوع 1-2', 'الفريق', ''],
            [3, 'إعداد مخطط التدفق', 'أسبوع 2-3', 'مدير الإنتاج', ''],
            [4, 'التحقق من مخطط التدفق', 'أسبوع 3', 'الفريق', ''],
            [5, 'تحليل المخاطر', 'أسبوع 3-5', 'الفريق', ''],
            [6, 'تحديد نقاط التحكم الحرجة', 'أسبوع 5-6', 'الفريق', ''],
            [7, 'وضع الحدود الحرجة', 'أسبوع 6-7', 'مدير الجودة', ''],
            [8, 'إنشاء نظام المراقبة', 'أسبوع 7-8', 'مدير الجودة', ''],
            [9, 'إجراءات التصحيح', 'أسبوع 8-9', 'مدير الجودة', ''],
            [10, 'إجراءات التحقق', 'أسبوع 9-10', 'مدير الجودة', ''],
            [11, 'نظام التوثيق', 'أسبوع 10-12', 'مدير الجودة', ''],
        ]),
        ('halal_certification.xlsx', 'شهادة الحلال الدولية', [
            [1, 'اختيار جهة الاعتماد', 'أسبوع 1', 'المدير العام', 'MUI, JAKIM, GCC'],
            [2, 'مراجعة المواد الخام', 'أسبوع 1-3', 'مدير الجودة', ''],
            [3, 'مراجعة خطوط الإنتاج', 'أسبوع 2-4', 'مدير الإنتاج', ''],
            [4, 'تدريب العاملين', 'أسبوع 3-5', 'الموارد البشرية', ''],
            [5, 'إعداد الوثائق', 'أسبوع 4-8', 'مدير الجودة', ''],
            [6, 'التدقيق الأولي', 'أسبوع 8-10', 'جهة الاعتماد', ''],
            [7, 'معالجة الملاحظات', 'أسبوع 10-12', 'الفريق', ''],
            [8, 'إصدار الشهادة', 'أسبوع 12', 'جهة الاعتماد', ''],
        ]),
        ('iso_9001_quality.xlsx', 'ISO 9001 - نظام إدارة الجودة', [
            [1, 'تحليل سياق المنظمة', 'أسبوع 1-2', 'المدير العام', ''],
            [2, 'تحديد الأطراف المعنية', 'أسبوع 2-3', 'الفريق', ''],
            [3, 'تحديد نطاق النظام', 'أسبوع 3', 'مدير الجودة', ''],
            [4, 'إعداد سياسة الجودة', 'أسبوع 3-4', 'المدير العام', ''],
            [5, 'تحديد المخاطر والفرص', 'أسبوع 4-6', 'الفريق', ''],
            [6, 'إعداد الأهداف', 'أسبوع 5-6', 'المدراء', ''],
            [7, 'توثيق العمليات', 'أسبوع 6-16', 'مدير الجودة', ''],
            [8, 'التطبيق', 'أسبوع 12-20', 'الجميع', ''],
            [9, 'التدقيق والشهادة', 'أسبوع 22-26', 'جهة الاعتماد', ''],
        ]),
        ('brc_global_standard.xlsx', 'BRC Global Standard - معيار سلامة الغذاء العالمي', [
            [1, 'فهم متطلبات BRC v9', 'أسبوع 1-4', 'مدير الجودة', ''],
            [2, 'تقييم الفجوة', 'أسبوع 4-8', 'استشاري خارجي', ''],
            [3, 'تطوير نظام الجودة', 'أسبوع 8-20', 'مدير الجودة', ''],
            [4, 'تحسين البنية التحتية', 'أسبوع 8-28', 'مدير المرافق', ''],
            [5, 'تدريب الموظفين', 'أسبوع 12-24', 'الموارد البشرية', ''],
            [6, 'المراجعة الداخلية', 'أسبوع 28-32', 'المراجع', ''],
            [7, 'تدقيق BRC', 'أسبوع 34-36', 'BRC', ''],
        ]),
        ('fssc_22000.xlsx', 'FSSC 22000 - شهادة نظام سلامة الغذاء', [
            [1, 'ISO 22000 كأساس', 'متطلب سابق', '-', 'يجب الحصول عليها أولاً'],
            [2, 'تطبيق ISO/TS 22002-1', 'أسبوع 1-12', 'مدير الجودة', ''],
            [3, 'متطلبات FSSC الإضافية', 'أسبوع 8-16', 'مدير الجودة', ''],
            [4, 'إدارة الغش الغذائي', 'أسبوع 10-14', 'مدير الجودة', ''],
            [5, 'دفاع الغذاء', 'أسبوع 12-16', 'مدير السلامة', ''],
            [6, 'تدقيق FSSC', 'أسبوع 20-24', 'جهة الاعتماد', ''],
        ]),
        ('organic_certification.xlsx', 'شهادة المنتجات العضوية', [
            [1, 'اختيار جهة الاعتماد', 'أسبوع 1', 'المدير العام', 'EU/USDA/JAS'],
            [2, 'تقييم سلسلة التوريد', 'أسبوع 1-4', 'المشتريات', ''],
            [3, 'فصل خطوط الإنتاج', 'أسبوع 4-12', 'مدير الإنتاج', ''],
            [4, 'نظام التتبع العضوي', 'أسبوع 8-16', 'مدير الجودة', ''],
            [5, 'توثيق الموردين', 'أسبوع 12-20', 'المشتريات', ''],
            [6, 'التدقيق والشهادة', 'أسبوع 24-28', 'جهة الاعتماد', ''],
        ]),
    ]

    headers = ['م', 'المهمة', 'الجدول', 'المسؤول', 'ملاحظات']

    for filename, title, data in certs:
        wb = Workbook()
        ws = wb.active
        ws.title = "خارطة الطريق"
        ws.sheet_view.rightToLeft = True
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:E1')

        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        style_header(ws, 3, len(headers))
        add_data(ws, 4, data)

        status_dv = DataValidation(type="list", formula1='"لم يبدأ,قيد التنفيذ,مكتمل"')
        ws.add_data_validation(status_dv)

        set_widths(ws, [5, 35, 15, 18, 30])
        wb.save(f'04_certifications/{filename}')
        print(f"    ✓ {filename}")

# ===================== 05_export_readiness =====================
def create_export():
    print("  جاري إنشاء ملفات جاهزية التصدير...")

    # متطلبات رخصة التصدير
    wb = Workbook()
    ws = wb.active
    ws.title = "متطلبات التصدير"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "متطلبات الحصول على رخصة التصدير السعودية"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    headers = ['م', 'المتطلب', 'الجهة', 'الوصف', 'الحالة', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    data = [
        [1, 'السجل التجاري', 'وزارة التجارة', 'ساري ويتضمن نشاط التصدير', '', ''],
        [2, 'شهادة الزكاة', 'ZATCA', 'سارية المفعول', '', ''],
        [3, 'التسجيل في نظام سابر', 'SASO', 'للمنتجات الصناعية', '', ''],
        [4, 'شهادة منشأ', 'الغرفة التجارية', 'لكل شحنة تصدير', '', ''],
        [5, 'شهادة صحية', 'SFDA', 'للمنتجات الغذائية', '', ''],
        [6, 'شهادة حلال', 'جهة معتمدة', 'للأسواق الإسلامية', '', ''],
        [7, 'التسجيل في منصة فسح', 'الجمارك', 'للتخليص الجمركي', '', ''],
    ]
    add_data(ws, 4, data)
    set_widths(ws, [5, 22, 18, 35, 12, 25])
    wb.save('05_export_readiness/export_license_requirements.xlsx')
    print("    ✓ export_license_requirements.xlsx")

    # تحليل الأسواق
    wb = Workbook()
    ws = wb.active
    ws.title = "الأسواق المستهدفة"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "تحليل الأسواق المستهدفة للتصدير"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:H1')

    headers = ['السوق', 'حجم السوق', 'النمو', 'المنافسة', 'المتطلبات', 'التعرفة', 'الأولوية', 'الحالة']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    markets = [
        ['الإمارات', '$15B', '8%', 'عالية', 'GSO + Emirates conformity', '0%', 'عالية', ''],
        ['الكويت', '$4B', '5%', 'متوسطة', 'GSO + KUCAS', '0%', 'عالية', ''],
        ['البحرين', '$2B', '4%', 'متوسطة', 'GSO', '0%', 'عالية', ''],
        ['قطر', '$5B', '6%', 'متوسطة', 'GSO + QS', '0%', 'عالية', ''],
        ['مصر', '$30B', '7%', 'عالية', 'NFSA + حلال', '5%', 'متوسطة', ''],
        ['الأردن', '$5B', '4%', 'متوسطة', 'JFDA', '0%', 'متوسطة', ''],
        ['المملكة المتحدة', '$100B', '3%', 'عالية جداً', 'UK CA + HACCP', '12%', 'متوسطة', ''],
        ['ألمانيا', '$200B', '2%', 'عالية جداً', 'EU + BRC/IFS', '12%', 'منخفضة', ''],
    ]
    add_data(ws, 4, markets)
    set_widths(ws, [15, 12, 8, 12, 30, 8, 12, 10])
    wb.save('05_export_readiness/target_markets_analysis.xlsx')
    print("    ✓ target_markets_analysis.xlsx")

    # متطلبات GCC
    wb = Workbook()
    ws = wb.active
    ws.title = "متطلبات الخليج"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "متطلبات التصدير لدول مجلس التعاون الخليجي"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')

    headers = ['م', 'المتطلب', 'التفاصيل', 'الإلزامية', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    data = [
        [1, 'مطابقة GSO', 'المواصفات القياسية الخليجية', 'إلزامي', 'أساسي'],
        [2, 'شهادة المنشأ', 'من الغرفة التجارية', 'إلزامي', 'لكل شحنة'],
        [3, 'شهادة صحية', 'من SFDA', 'إلزامي', 'للأغذية'],
        [4, 'شهادة حلال', 'من جهة معتمدة', 'إلزامي', 'للأغذية'],
        [5, 'فاتورة تجارية', 'موثقة من الغرفة', 'إلزامي', ''],
        [6, 'قائمة التعبئة', 'مفصلة', 'إلزامي', ''],
        [7, 'بوليصة الشحن', 'من شركة الشحن', 'إلزامي', ''],
        [8, 'شهادة تحليل', 'من مختبر معتمد', 'حسب المنتج', ''],
    ]
    add_data(ws, 4, data)
    set_widths(ws, [5, 20, 30, 12, 25])
    wb.save('05_export_readiness/gcc_requirements.xlsx')
    print("    ✓ gcc_requirements.xlsx")

    # EU, USA, Labeling, Customs files (simplified)
    for filename, title in [
        ('eu_requirements.xlsx', 'متطلبات التصدير للاتحاد الأوروبي'),
        ('usa_fda_requirements.xlsx', 'متطلبات FDA الأمريكية'),
        ('labeling_standards.xlsx', 'معايير البطاقة الغذائية الدولية'),
        ('customs_procedures.xlsx', 'إجراءات التخليص الجمركي'),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "المتطلبات"
        ws.sheet_view.rightToLeft = True
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:E1')

        headers = ['م', 'المتطلب', 'التفاصيل', 'الإلزامية', 'الحالة']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        style_header(ws, 3, len(headers))
        set_widths(ws, [5, 25, 40, 12, 12])
        wb.save(f'05_export_readiness/{filename}')
        print(f"    ✓ {filename}")

# ===================== 06_risk_management =====================
def create_risk_mgmt():
    print("  جاري إنشاء ملفات إدارة المخاطر...")

    # سجل المخاطر
    wb = Workbook()
    ws = wb.active
    ws.title = "سجل المخاطر"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "سجل المخاطر الشامل"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:J1')

    headers = ['م', 'المخاطرة', 'الفئة', 'الاحتمالية', 'الأثر', 'الدرجة', 'إجراء التخفيف', 'المسؤول', 'الحالة', 'تاريخ المراجعة']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    risks = [
        [1, 'تلوث المنتج', 'سلامة غذاء', 4, 5, '=D4*E4', 'نظام HACCP + فحوصات', 'مدير الجودة', '', ''],
        [2, 'انقطاع المورد', 'سلسلة توريد', 3, 4, '=D5*E5', 'موردين بدلاء', 'المشتريات', '', ''],
        [3, 'عدم الامتثال التنظيمي', 'امتثال', 3, 5, '=D6*E6', 'نظام مراقبة', 'مدير الامتثال', '', ''],
        [4, 'حريق', 'تشغيلي', 2, 5, '=D7*E7', 'نظام إطفاء + تأمين', 'مدير السلامة', '', ''],
        [5, 'تسريب بيانات', 'أمن معلومات', 2, 3, '=D8*E8', 'نظام حماية', 'IT', '', ''],
        [6, 'استدعاء منتج', 'سمعة', 2, 5, '=D9*E9', 'نظام تتبع + إجراء استدعاء', 'مدير الجودة', '', ''],
        [7, 'إصابة عمل', 'صحة وسلامة', 3, 4, '=D10*E10', 'تدريب + معدات حماية', 'مدير السلامة', '', ''],
        [8, 'انقطاع الكهرباء', 'تشغيلي', 2, 3, '=D11*E11', 'مولد احتياطي', 'مدير المرافق', '', ''],
    ]
    add_data(ws, 4, risks)

    # Data Validation
    prob_dv = DataValidation(type="list", formula1='"1,2,3,4,5"')
    ws.add_data_validation(prob_dv)
    prob_dv.add('D4:D100')
    prob_dv.add('E4:E100')

    status_dv = DataValidation(type="list", formula1='"مفتوح,قيد المعالجة,مغلق,مقبول"')
    ws.add_data_validation(status_dv)
    status_dv.add('I4:I100')

    # Conditional formatting for risk score
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
    green_fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")

    ws.conditional_formatting.add('F4:F100', FormulaRule(formula=['$F4>=15'], fill=red_fill))
    ws.conditional_formatting.add('F4:F100', FormulaRule(formula=['AND($F4>=8,$F4<15)'], fill=yellow_fill))
    ws.conditional_formatting.add('F4:F100', FormulaRule(formula=['$F4<8'], fill=green_fill))

    set_widths(ws, [5, 25, 15, 12, 10, 10, 30, 15, 12, 15])
    wb.save('06_risk_management/risk_register.xlsx')
    print("    ✓ risk_register.xlsx")

    # حاسبة الغرامات
    wb = Workbook()
    ws = wb.active
    ws.title = "حاسبة الغرامات"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "حاسبة الغرامات والعقوبات"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')

    headers = ['المخالفة', 'الجهة', 'الغرامة الأولى', 'التكرار', 'الحد الأقصى', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    penalties = [
        ['منتج منتهي الصلاحية', 'SFDA', 100000, 500000, 1000000, 'قد يصل للسجن'],
        ['عدم توفر HACCP', 'SFDA', 50000, 150000, 200000, ''],
        ['عدم توثيق العقود', 'HRSD', 10000, 20000, 100000, 'لكل عامل'],
        ['تأخير الرواتب', 'HRSD', 10000, 50000, 100000, 'إيقاف خدمات'],
        ['عدم سداد VAT', 'ZATCA', '5%', '25%', '50%', 'من المبلغ'],
        ['عدم توفر طفايات', 'الدفاع المدني', 20000, 50000, 100000, 'إغلاق'],
    ]
    add_data(ws, 4, penalties)
    set_widths(ws, [25, 18, 15, 15, 15, 25])
    wb.save('06_risk_management/penalties_calculator.xlsx')
    print("    ✓ penalties_calculator.xlsx")

    # خطط الطوارئ
    wb = Workbook()
    ws = wb.active
    ws.title = "خطط الطوارئ"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "خطط الطوارئ والاستجابة"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')

    headers = ['السيناريو', 'المؤشرات', 'الاستجابة الفورية', 'المسؤول', 'التصعيد']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    data = [
        ['استدعاء منتج', 'شكاوى + نتائج تحليل', 'إيقاف التوزيع + إخطار SFDA', 'مدير الجودة', 'المدير العام'],
        ['حريق', 'إنذار + دخان', 'إخلاء + اتصال 998', 'مدير السلامة', 'المدير العام'],
        ['تسمم غذائي', 'بلاغات + أعراض', 'عزل المنتج + تحقيق', 'مدير الجودة', 'SFDA'],
        ['انقطاع كهرباء', 'أجهزة الإنذار', 'تشغيل المولد + فحص المبردات', 'مدير المرافق', 'مدير الإنتاج'],
        ['تسرب مواد', 'رائحة + بقع', 'إخلاء + احتواء', 'مدير السلامة', 'الدفاع المدني'],
    ]
    add_data(ws, 4, data)
    set_widths(ws, [18, 25, 35, 18, 18])
    wb.save('06_risk_management/contingency_plans.xlsx')
    print("    ✓ contingency_plans.xlsx")

    # متطلبات التأمين
    wb = Workbook()
    ws = wb.active
    ws.title = "التأمين"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "متطلبات التأمين"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')

    headers = ['نوع التأمين', 'التغطية', 'الحد الأدنى', 'الإلزامية', 'الشركة']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    data = [
        ['تأمين المنشأة', 'حريق + سرقة + كوارث', '5 مليون', 'إلزامي', ''],
        ['تأمين المسؤولية', 'المنتجات', '10 مليون', 'موصى به', ''],
        ['تأمين العمال', 'إصابات العمل', 'حسب العدد', 'إلزامي', 'التأمينات'],
        ['تأمين المركبات', 'الشاحنات', 'شامل', 'إلزامي', ''],
        ['تأمين البضائع', 'أثناء النقل', 'قيمة البضاعة', 'موصى به', ''],
    ]
    add_data(ws, 4, data)
    set_widths(ws, [20, 25, 15, 12, 20])
    wb.save('06_risk_management/insurance_requirements.xlsx')
    print("    ✓ insurance_requirements.xlsx")

# ===================== 07_kpis_dashboard =====================
def create_kpis():
    print("  جاري إنشاء لوحة المؤشرات...")

    # بطاقة الأداء
    wb = Workbook()
    ws = wb.active
    ws.title = "بطاقة الأداء"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "بطاقة أداء الامتثال"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    headers = ['المؤشر', 'الهدف', 'الفعلي', 'النسبة %', 'الاتجاه', 'الحالة', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    kpis = [
        ['التراخيص السارية', '100%', '', '=IF(C4="","",C4/B4*100)', '', '', ''],
        ['تطبيق HACCP', '100%', '', '=IF(C5="","",C5/B5*100)', '', '', ''],
        ['تدريب الموظفين', '100%', '', '=IF(C6="","",C6/B6*100)', '', '', ''],
        ['مخالفات قائمة', '0', '', '=IF(C7="","",(B7-C7)/B7*100)', '', '', ''],
        ['نتائج التدقيق', '90%', '', '=IF(C8="","",C8)', '', '', ''],
        ['شكاوى العملاء', '<5/شهر', '', '', '', '', ''],
        ['استدعاءات المنتج', '0', '', '', '', '', ''],
        ['إصابات العمل', '0', '', '', '', '', ''],
    ]
    add_data(ws, 4, kpis)

    status_dv = DataValidation(type="list", formula1='"أخضر,أصفر,أحمر"')
    ws.add_data_validation(status_dv)
    status_dv.add('F4:F100')

    trend_dv = DataValidation(type="list", formula1='"↑,↓,→"')
    ws.add_data_validation(trend_dv)
    trend_dv.add('E4:E100')

    set_widths(ws, [22, 12, 12, 12, 10, 10, 25])
    wb.save('07_kpis_dashboard/compliance_scorecard.xlsx')
    print("    ✓ compliance_scorecard.xlsx")

    # باقي ملفات KPIs
    for filename, title in [
        ('monthly_tracking.xlsx', 'المتابعة الشهرية'),
        ('audit_schedule.xlsx', 'جدول المراجعات والتدقيق'),
        ('management_reports.xlsx', 'تقارير الإدارة'),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "البيانات"
        ws.sheet_view.rightToLeft = True
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:F1')

        headers = ['م', 'البند', 'التاريخ', 'المسؤول', 'الحالة', 'ملاحظات']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        style_header(ws, 3, len(headers))
        set_widths(ws, [5, 30, 15, 18, 12, 25])
        wb.save(f'07_kpis_dashboard/{filename}')
        print(f"    ✓ {filename}")

# ===================== 08_documentation =====================
def create_documentation():
    print("  جاري إنشاء ملفات التوثيق...")

    # فهرس السياسات
    wb = Workbook()
    ws = wb.active
    ws.title = "السياسات"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "فهرس السياسات والإجراءات"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    headers = ['الرقم', 'الوثيقة', 'النوع', 'الإصدار', 'تاريخ الإصدار', 'المراجعة القادمة', 'الحالة']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    docs = [
        ['POL-001', 'سياسة سلامة الغذاء', 'سياسة', '1.0', '', '', ''],
        ['POL-002', 'سياسة الجودة', 'سياسة', '1.0', '', '', ''],
        ['POL-003', 'سياسة الصحة والسلامة', 'سياسة', '1.0', '', '', ''],
        ['SOP-001', 'إجراء استلام المواد الخام', 'إجراء', '1.0', '', '', ''],
        ['SOP-002', 'إجراء التخزين', 'إجراء', '1.0', '', '', ''],
        ['SOP-003', 'إجراء الإنتاج', 'إجراء', '1.0', '', '', ''],
        ['SOP-004', 'إجراء التعبئة والتغليف', 'إجراء', '1.0', '', '', ''],
        ['SOP-005', 'إجراء التنظيف والتعقيم', 'إجراء', '1.0', '', '', ''],
        ['WI-001', 'تعليمات غسل اليدين', 'تعليمات', '1.0', '', '', ''],
        ['FOR-001', 'نموذج سجل درجات الحرارة', 'نموذج', '1.0', '', '', ''],
    ]
    add_data(ws, 4, docs)

    status_dv = DataValidation(type="list", formula1='"معتمد,مسودة,قيد المراجعة,ملغي"')
    ws.add_data_validation(status_dv)
    status_dv.add('G4:G100')

    type_dv = DataValidation(type="list", formula1='"سياسة,إجراء,تعليمات,نموذج,دليل"')
    ws.add_data_validation(type_dv)
    type_dv.add('C4:C100')

    set_widths(ws, [12, 30, 12, 10, 15, 15, 12])
    wb.save('08_documentation/policies_procedures.xlsx')
    print("    ✓ policies_procedures.xlsx")

    # باقي ملفات التوثيق
    for filename, title in [
        ('records_retention.xlsx', 'جدول حفظ السجلات'),
        ('training_matrix.xlsx', 'مصفوفة التدريب'),
        ('document_control.xlsx', 'سجل ضبط الوثائق'),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "البيانات"
        ws.sheet_view.rightToLeft = True
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:F1')

        headers = ['م', 'البند', 'التفاصيل', 'المدة', 'المسؤول', 'ملاحظات']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        style_header(ws, 3, len(headers))
        set_widths(ws, [5, 25, 30, 12, 18, 25])
        wb.save(f'08_documentation/{filename}')
        print(f"    ✓ {filename}")

# ===================== 09_financial_planning =====================
def create_financial():
    print("  جاري إنشاء ملفات التخطيط المالي...")

    # ميزانية الامتثال
    wb = Workbook()
    ws = wb.active
    ws.title = "الميزانية"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "ميزانية مشروع الامتثال"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    headers = ['البند', 'التفاصيل', 'الربع 1', 'الربع 2', 'الربع 3', 'الربع 4', 'الإجمالي']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    budget = [
        ['التراخيص والتجديدات', 'السجل + SFDA + البلدية', 50000, 10000, 10000, 10000, '=SUM(C4:F4)'],
        ['الاستشارات', 'استشاري HACCP + ISO', 40000, 40000, 20000, 0, '=SUM(C5:F5)'],
        ['الشهادات الدولية', 'ISO + HACCP + حلال', 0, 50000, 80000, 50000, '=SUM(C6:F6)'],
        ['التدريب', 'تدريب الموظفين', 20000, 15000, 15000, 10000, '=SUM(C7:F7)'],
        ['المعدات والبنية التحتية', 'تحديثات المصنع', 100000, 80000, 50000, 20000, '=SUM(C8:F8)'],
        ['المختبرات والتحاليل', 'تحاليل دورية', 10000, 10000, 10000, 10000, '=SUM(C9:F9)'],
        ['طوارئ 10%', 'احتياطي', 22000, 20500, 18500, 10000, '=SUM(C10:F10)'],
        ['الإجمالي', '', '=SUM(C4:C10)', '=SUM(D4:D10)', '=SUM(E4:E10)', '=SUM(F4:F10)', '=SUM(G4:G10)'],
    ]
    add_data(ws, 4, budget)

    # تنسيق صف الإجمالي
    for col in range(1, 8):
        ws.cell(row=11, column=col).font = Font(bold=True)

    set_widths(ws, [28, 25, 15, 15, 15, 15, 15])
    wb.save('09_financial_planning/compliance_budget.xlsx')
    print("    ✓ compliance_budget.xlsx")

    # باقي الملفات
    for filename, title in [
        ('roi_analysis.xlsx', 'تحليل العائد على الاستثمار'),
        ('funding_options.xlsx', 'خيارات التمويل المتاحة'),
        ('incentives_subsidies.xlsx', 'الحوافز والدعم الحكومي'),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "البيانات"
        ws.sheet_view.rightToLeft = True
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:E1')

        headers = ['م', 'البند', 'القيمة', 'الملاحظات', 'المصدر']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        style_header(ws, 3, len(headers))
        set_widths(ws, [5, 30, 20, 30, 20])
        wb.save(f'09_financial_planning/{filename}')
        print(f"    ✓ {filename}")

# ===================== 10_checklists =====================
def create_checklists():
    print("  جاري إنشاء قوائم الفحص...")

    # قائمة فحص SFDA
    wb = Workbook()
    ws = wb.active
    ws.title = "قائمة الفحص"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "قائمة فحص هيئة الغذاء والدواء"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:E1')

    headers = ['م', 'البند', 'مطابق', 'غير مطابق', 'ملاحظات']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    items = [
        [1, 'الترخيص ساري ومعلق في مكان واضح', '☐', '☐', ''],
        [2, 'وجود نظام HACCP موثق', '☐', '☐', ''],
        [3, 'شهادات صحية سارية لجميع العاملين', '☐', '☐', ''],
        [4, 'سجلات درجات الحرارة محدثة', '☐', '☐', ''],
        [5, 'نظافة عامة مقبولة', '☐', '☐', ''],
        [6, 'عدم وجود آفات', '☐', '☐', ''],
        [7, 'المنتجات مؤرخة وصالحة', '☐', '☐', ''],
        [8, 'التخزين حسب الشروط', '☐', '☐', ''],
        [9, 'البطاقة الغذائية مطابقة', '☐', '☐', ''],
        [10, 'نظام التتبع فعال', '☐', '☐', ''],
        [11, 'معدات الحماية متوفرة', '☐', '☐', ''],
        [12, 'مرافق غسل اليدين كافية', '☐', '☐', ''],
        [13, 'فصل بين المناطق', '☐', '☐', ''],
        [14, 'إضاءة كافية', '☐', '☐', ''],
        [15, 'تهوية مناسبة', '☐', '☐', ''],
    ]
    add_data(ws, 4, items)
    set_widths(ws, [5, 40, 12, 12, 30])
    wb.save('10_checklists/sfda_inspection_checklist.xlsx')
    print("    ✓ sfda_inspection_checklist.xlsx")

    # باقي القوائم
    for filename, title in [
        ('pre_audit_checklist.xlsx', 'قائمة ما قبل التدقيق'),
        ('export_readiness_checklist.xlsx', 'قائمة جاهزية التصدير'),
        ('daily_compliance_checklist.xlsx', 'قائمة الامتثال اليومي'),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "القائمة"
        ws.sheet_view.rightToLeft = True
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:E1')

        headers = ['م', 'البند', 'نعم', 'لا', 'ملاحظات']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        style_header(ws, 3, len(headers))
        set_widths(ws, [5, 45, 10, 10, 30])
        wb.save(f'10_checklists/{filename}')
        print(f"    ✓ {filename}")

# ===================== 11_master_files =====================
def create_master():
    print("  جاري إنشاء الملفات الرئيسية...")

    # ميثاق المشروع
    wb = Workbook()
    ws = wb.active
    ws.title = "ميثاق المشروع"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "ميثاق مشروع الامتثال"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:D1')

    charter = [
        ['اسم المشروع', 'مشروع امتثال مصنع الأغذية للمتطلبات النظامية'],
        ['راعي المشروع', '[اسم المدير العام]'],
        ['مدير المشروع', '[اسم مدير المشروع]'],
        ['تاريخ البداية', '[التاريخ]'],
        ['تاريخ الانتهاء المتوقع', '[التاريخ]'],
        ['الميزانية', '[المبلغ]'],
        ['', ''],
        ['الأهداف الرئيسية', ''],
        ['1', 'رفع إيقاف المنشأة خلال 30 يوم'],
        ['2', 'الحصول على جميع التراخيص خلال 90 يوم'],
        ['3', 'تطبيق نظام HACCP خلال 180 يوم'],
        ['4', 'الحصول على ISO 22000 خلال 12 شهر'],
        ['5', 'الجاهزية للتصدير خلال 18 شهر'],
    ]

    for row_idx, (col1, col2) in enumerate(charter, 3):
        ws.cell(row=row_idx, column=1, value=col1).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = RTL_ALIGN
        ws.cell(row=row_idx, column=2, value=col2)
        ws.cell(row=row_idx, column=2).alignment = RTL_ALIGN

    set_widths(ws, [25, 50])
    wb.save('11_master_files/project_charter.xlsx')
    print("    ✓ project_charter.xlsx")

    # سجل أصحاب المصلحة
    wb = Workbook()
    ws = wb.active
    ws.title = "أصحاب المصلحة"
    ws.sheet_view.rightToLeft = True
    ws['A1'] = "سجل أصحاب المصلحة"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:G1')

    headers = ['م', 'صاحب المصلحة', 'الدور', 'التأثير', 'الاهتمام', 'استراتيجية التواصل', 'التردد']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))

    stakeholders = [
        [1, 'المدير العام', 'راعي المشروع', 'عالي', 'عالي', 'تقارير مباشرة', 'أسبوعي'],
        [2, 'هيئة الغذاء والدواء', 'جهة رقابية', 'عالي', 'عالي', 'تقارير رسمية', 'حسب الطلب'],
        [3, 'وزارة التجارة', 'جهة ترخيص', 'عالي', 'متوسط', 'طلبات رسمية', 'حسب الحاجة'],
        [4, 'الموظفون', 'منفذون', 'متوسط', 'عالي', 'اجتماعات + تدريب', 'شهري'],
        [5, 'العملاء', 'مستفيدون', 'متوسط', 'متوسط', 'إعلانات', 'ربع سنوي'],
        [6, 'الموردون', 'شركاء', 'متوسط', 'متوسط', 'اجتماعات', 'ربع سنوي'],
    ]
    add_data(ws, 4, stakeholders)
    set_widths(ws, [5, 22, 18, 12, 12, 25, 12])
    wb.save('11_master_files/stakeholder_register.xlsx')
    print("    ✓ stakeholder_register.xlsx")

    # الملفات المتبقية
    for filename, title in [
        ('communication_plan.xlsx', 'خطة التواصل'),
        ('master_timeline.xlsx', 'الجدول الزمني الرئيسي'),
        ('success_criteria.xlsx', 'معايير النجاح'),
    ]:
        wb = Workbook()
        ws = wb.active
        ws.title = "البيانات"
        ws.sheet_view.rightToLeft = True
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws.merge_cells('A1:F1')

        headers = ['م', 'البند', 'التفاصيل', 'المسؤول', 'الموعد', 'الحالة']
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        style_header(ws, 3, len(headers))
        set_widths(ws, [5, 25, 35, 18, 15, 12])
        wb.save(f'11_master_files/{filename}')
        print(f"    ✓ {filename}")

# ===================== التنفيذ الرئيسي =====================
if __name__ == "__main__":
    os.chdir('/home/user/chtgpt/food_factory_compliance')
    print("\n" + "="*60)
    print("جاري إنشاء جميع ملفات مشروع الامتثال...")
    print("="*60 + "\n")

    create_roadmap()
    create_certifications()
    create_export()
    create_risk_mgmt()
    create_kpis()
    create_documentation()
    create_financial()
    create_checklists()
    create_master()

    print("\n" + "="*60)
    print("✓ تم إنشاء جميع الملفات بنجاح!")
    print("="*60)
